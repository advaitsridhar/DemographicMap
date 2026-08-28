#!/usr/bin/env python3
"""What is actually in a published workbook: its sheets, and their headers.

A spreadsheet is worth an adapter only if it carries the columns the map
needs, and a title never settles that. Bangladesh's census release is called
"Population and Housing Census dataset ... Admin 02", which says the geography
and says nothing about whether religion is among the several hundred
indicators beside it.

Reading it costs one download. The alternative is writing an adapter against a
guess and finding out from the failure, which costs a run and teaches less.

Sheets are opened read-only and streamed, so a workbook far larger than memory
still reports. A sheet is described by its first rows because that is where a
statistical office puts the header, and by its dimensions because a sheet with
four columns and one with four hundred are different problems.

Read-only, and the output is the log.

Usage:
    python scripts/probe_xlsx.py <url> --rows 3 --match religion,muslim
"""

from __future__ import annotations

import argparse
import io
import sys
import urllib.request

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; DemographicMap/1.0; "
                  "+https://github.com/advaitsridhar/DemographicMap)",
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
}
TIMEOUT = 120


def log(message: str = "") -> None:
    print(message, flush=True)


def fetch(url: str) -> bytes:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        return resp.read()


def cells(row) -> list[str]:
    """One row as text, with the trailing empties dropped.

    A sheet's declared width is the widest row in it, so a header read without
    trimming ends in a run of Nones that says nothing and hides what does.
    """
    out = ["" if value is None else str(value).strip() for value in row]
    while out and not out[-1]:
        out.pop()
    return out


def describe(blob: bytes, rows: int, match: list[str], width: int,
             wanted: list[str], cols: int) -> None:
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    log(f"  {len(book.sheetnames)} sheet(s): {', '.join(book.sheetnames)}")
    for name in book.sheetnames:
        if wanted and not any(term in name.strip().lower() for term in wanted):
            continue
        sheet = book[name]
        log(f"\n--- sheet {name!r} "
            f"({sheet.max_row or '?'} rows x {sheet.max_column or '?'} cols) "
            + "-" * 20)
        seen: list[list[str]] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= rows:
                break
            seen.append(cells(row))
        for line in seen:
            log("  " + " | ".join(value[:width] for value in line[:cols]))
        if match:
            # Searched across every row read, not just the first: these
            # releases often carry a title row, a blank, then the header.
            hits = sorted({value for line in seen for value in line
                           if any(term in value.lower() for term in match)})
            log(f"  matching {match}: {hits if hits else 'none in these rows'}")
    book.close()


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("url", nargs="+")
    ap.add_argument("--rows", type=int, default=3,
                    help="rows to print from each sheet")
    ap.add_argument("--match", default="",
                    help="comma-separated substrings to report from those rows")
    ap.add_argument("--width", type=int, default=40,
                    help="characters per cell")
    ap.add_argument("--sheet", default="",
                    help="comma-separated substrings; only these sheets")
    ap.add_argument("--cols", type=int, default=60,
                    help="columns to print from each row")
    args = ap.parse_args()

    match = [t.strip().lower() for t in args.match.split(",") if t.strip()]
    wanted = [t.strip().lower() for t in args.sheet.split(",") if t.strip()]
    for url in args.url:
        log(f"probe_xlsx: {url}")
        try:
            blob = fetch(url)
        except Exception as err:                  # noqa: BLE001
            # The whole message: a TLS failure names the host the certificate
            # is valid for, and that name is the finding.
            log(f"  unreachable: {type(err).__name__}: {str(err)[:400]}")
            continue
        log(f"  {len(blob):,} bytes")
        try:
            describe(blob, args.rows, match, args.width, wanted, args.cols)
        except Exception as err:                  # noqa: BLE001
            # A soft 404 is an HTML page with a spreadsheet's name, and it
            # fails here rather than at the fetch.
            log(f"  not a workbook: {type(err).__name__}: {str(err)[:200]}")
        log("")
    return 0


if __name__ == "__main__":
    sys.exit(main())
