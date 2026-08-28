#!/usr/bin/env python3
"""Bangladesh -- Population and Housing Census 2022, religion by zila.

The Bureau of Statistics publishes a workbook of census indicators at admin-2,
one row per zila, and among its forty-two sheets is *Population by Religion,
Sex*: Muslim, Hindu, Christian, Buddhist and Others, each as a total and by
sex, for all sixty-four districts.

**It is read from a mirror, and that is a deliberate choice rather than a
convenience.** None of the office's own hosts can be fetched over a connection
that verifies:

* ``bbs.gov.bd`` has a valid Sectigo certificate that does cover the host, but
  the server never sends its intermediate, so no chain can be built. A browser
  papers over this by fetching the issuer named in the certificate; urllib does
  not.
* ``bbs.portal.gov.bd`` answers with a "Kubernetes Ingress Controller Fake
  Certificate" for ``ingress.local``.
* ``file.portal.gov.bd``, ``sid.portal.gov.bd`` and ``portal.gov.bd`` time out.

The alternative to a mirror is disabling certificate verification, which this
project does not do, or recording Bangladesh as uncollectable, which would be
false -- the census exists, is published, and is CC0. HDX carries the release
under the UN in Bangladesh, and the workbook is the office's own.

**The religion table does not count everybody.** Each religion's total is
exactly its male plus its female column, and Bangladesh enumerates a third
gender: Barguna's religions sum to 1,010,461 against a district population of
1,010,531, and the 70 missing are its hijra -- the number the population sheet
prints for Barguna in its own Hijra column. So the shares here are of the
population the religion table classifies, and the note says so. Both figures
are kept -- the district's own population, and the denominator the shares are
of -- because silently using one for the other is how a footnote becomes a
wrong number.

That also makes the check exact rather than approximate: the religions plus
the hijra must equal the published total, to the person, in every district.

**The workbook's own merged sheet is not used, because it is wrong.**
``Merged_All_Table`` flattens the forty-two sheets into 445 columns, and in it
Cumilla and Cox's Bazar hold each other's household and population figures --
Cumilla 2,823,268 against a real 6.2 million -- while their district geocodes,
19 and 22, stay correct. Joypurhat and Naogaon are wrong too, and Naogaon's
figure matches neither district, so it is not a clean transposition throughout.
The per-topic sheets it was built from are consistent, and those are read
instead. Nothing here takes the division names from the merged sheet either:
they may well be sound, but a sheet with three known transpositions in it is
not something to take an unverifiable field from.

Usage:
    python -m scripts.fetch_census.bangladesh
"""

from __future__ import annotations

import argparse
import io
from typing import Any

from ._shared import (
    NOT_AVAILABLE, PROCESSED, gap, log, measure, record, shares, write_json,
)

SOURCE = ("Bangladesh Bureau of Statistics, Population and Housing Census 2022, "
          "district-level indicators")
URL = "https://data.humdata.org/dataset/populationa-and-housing-census-dataset"
LICENCE = "CC0 (public domain), published via HDX by the UN in Bangladesh"
YEAR = 2022

WORKBOOK = ("https://data.humdata.org/dataset/"
            "a6fedebe-72fe-4fc2-8657-1580acfa32c6/resource/"
            "72eaaa6c-6a30-4efd-bad9-02133b316ea8/download/"
            "bangladesh_bbs_population-and-housing-census-dataset_2022_admin-02.xlsx")

# Matched after stripping: the religion sheet's name begins with a space in
# the published file, and a lookup by the name as it reads would miss it.
RELIGION_SHEET = "Population by Religion, Sex"
POPULATION_SHEET = "Population by Sex, Dist & Loca"

# The order they appear in, and the names this map uses for them. "Others" is
# the office's own residual and is kept as one, rather than being dropped or
# guessed at.
RELIGIONS = ["Muslim", "Hindu", "Christian", "Buddhist", "Other religion"]
COLUMNS = {"Muslim": "# Total_Muslim", "Hindu": "# Total_Hindu",
           "Christian": "# Total_Christian", "Buddhist": "# Total_Buddhist",
           "Other religion": "# Total_Others"}
SEXED = {"Muslim": ("# Male_Muslim", "# Female_Muslim"),
         "Hindu": ("# Male_Hindu", "# Female_Hindu"),
         "Christian": ("# Male_Christian", "# Female_Christian"),
         "Buddhist": ("# Male_Buddhist", "# Female_Buddhist"),
         "Other religion": ("# Male_Others", "# Female_Others")}

# What geoBoundaries calls the same zila. Bangladesh respelled several
# districts in English in 2018 -- Chittagong became Chattogram, Comilla became
# Cumilla, Barisal Barishal, Jessore Jashore, Bogra Bogura -- and the boundary
# file still carries the older forms, alongside plain transliteration variants
# for three more. Declared rather than derived: "Nawabganj" and
# "Chapainababganj" share no word, and a rule loose enough to bridge them would
# bridge a great deal else.
ALIASES: dict[str, tuple[str, ...]] = {
    "Barishal": ("Barisal",),
    "Bogura": ("Bogra",),
    "Brahmanbaria": ("Brahamanbaria",),
    "Chapainababganj": ("Nawabganj", "Chapai Nawabganj"),
    "Chattogram": ("Chittagong",),
    "Cumilla": ("Comilla",),
    "Jashore": ("Jessore",),
    "Moulvibazar": ("Maulvibazar",),
}

NOTE = ("Census 2022. The religion table classifies the male and female "
        "population only -- each religion's total is exactly its male plus "
        "its female column -- so these shares are of a denominator a few "
        "dozen people short of the district's own population, the difference "
        "being the third-gender (hijra) population the table does not "
        "classify.")


def fetch(url: str) -> bytes:
    import urllib.request

    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (compatible; DemographicMap/1.0; "
                      "+https://github.com/advaitsridhar/DemographicMap)",
        "Accept": "application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet,*/*",
    })
    with urllib.request.urlopen(req, timeout=180) as resp:
        return resp.read()


def sheet(book, wanted: str):
    """A sheet by its name with the surrounding space ignored.

    The published workbook names one of them " Population by Religion, Sex",
    with a leading space, and a lookup by the name as it reads would report a
    sheet that is plainly there as missing.
    """
    for name in book.sheetnames:
        if name.strip() == wanted:
            return book[name]
    raise SystemExit(
        f"no sheet named {wanted!r} in the workbook; it has: "
        + ", ".join(repr(n) for n in book.sheetnames))


def table(sheet) -> list[dict[str, Any]]:
    """A sheet's rows as dicts, stopping where the data does.

    openpyxl reports the declared dimension rather than the used one -- these
    sheets say a thousand rows and hold sixty-four -- so the end of the table
    is the first row with no district on it, not the end of the sheet.
    """
    rows = sheet.iter_rows(values_only=True)
    header = [("" if cell is None else str(cell).strip())
              for cell in next(rows)]
    out = []
    for values in rows:
        if not values or values[0] is None or not str(values[0]).strip():
            break
        out.append(dict(zip(header, values)))
    return out


def pick(row: dict[str, Any], prefix: str) -> Any:
    """One column by its name, or by the only name that starts with it.

    The population sheet's headings run "Population_Total",
    "Population_Hijra", "Population_rural_..." and so on, and reading a
    heading off a printed excerpt truncates it. A prefix that matches exactly
    one column is the column; a prefix that matches several is ambiguous and
    says so rather than taking the first.
    """
    if prefix in row:
        return row[prefix]
    hits = [key for key in row if key.startswith(prefix)]
    if len(hits) == 1:
        return row[hits[0]]
    raise SystemExit(
        f"{prefix!r} matches {len(hits)} columns ({hits[:4]}) of: "
        + ", ".join(list(row)[:14]))


def number(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace(",", "").strip()
    try:
        return int(float(text))
    except ValueError:
        return None


def check(districts: list[dict[str, Any]]) -> None:
    """The workbook's own arithmetic, checked before any of it is used.

    Each religion's total must be its male plus its female column: that is the
    sheet stating the same figure twice, and the two agreeing is what makes
    the column headings trustworthy rather than assumed.

    And the religions plus the third gender must equal the district's
    published population exactly. Two separate sheets have to agree to the
    person for that to hold, which is what makes it worth having: the
    workbook's merged sheet fails this badly enough to have swapped two
    districts' populations, and a tolerance wide enough to admit a hijra count
    would have been wide enough to hide something worse.
    """
    bad = []
    for row in districts:
        for religion, (male, female) in SEXED.items():
            total = row["counts"][religion]
            parts = row["sexed"][religion]
            if None in parts:
                bad.append(f"{row['name']}: {religion} has no {male}/{female}")
            elif sum(parts) != total:
                bad.append(f"{row['name']}: {religion} totals {total:,} "
                           f"against {sum(parts):,} by sex")
        classified = sum(row["counts"].values())
        whole, hijra = row["population"], row["hijra"]
        if whole is None or hijra is None:
            bad.append(f"{row['name']}: no published population or hijra count")
        elif classified + hijra != whole:
            bad.append(f"{row['name']}: {classified:,} classified by religion "
                       f"plus {hijra:,} hijra is {classified + hijra:,}, "
                       f"against a published {whole:,}")
    if bad:
        raise SystemExit(f"{len(bad)} checks failed — " + "; ".join(bad[:4]))
    hijra = sum(row["hijra"] for row in districts)
    log(f"    every religion's total matches its own male plus female, and "
        f"in every district the religions plus the hijra come to the "
        f"published population exactly ({hijra:,} hijra nationally, whom the "
        f"religion table does not classify)")


def read(blob: bytes) -> list[dict[str, Any]]:
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        people = table(sheet(book, POPULATION_SHEET))
        religion = table(sheet(book, RELIGION_SHEET))
    finally:
        book.close()
    log(f"    {len(people)} districts in {POPULATION_SHEET}, "
        f"{len(religion)} in {RELIGION_SHEET}")

    whole = {str(row["District"]).strip(): row for row in people}
    missing = sorted({str(r["District"]).strip() for r in religion} - set(whole))
    if missing:
        # Named, because a district in one sheet and not the other is a fact
        # about the workbook rather than a row to quietly drop.
        raise SystemExit("districts in the religion sheet and not in "
                         f"{POPULATION_SHEET}: {', '.join(missing)}")

    out = []
    for row in religion:
        name = str(row["District"]).strip()
        counts = {religion_name: number(row.get(column))
                  for religion_name, column in COLUMNS.items()}
        if any(value is None for value in counts.values()):
            raise SystemExit(f"{name}: a religion column is missing or not a "
                             f"number: {counts}")
        out.append({
            "name": name,
            "population": number(pick(whole[name], "Population_Total")),
            "hijra": number(pick(whole[name], "Population_Hijra")),
            "counts": counts,
            "sexed": {religion_name: tuple(number(row.get(column))
                                           for column in pair)
                      for religion_name, pair in SEXED.items()},
        })
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    log("bangladesh: BBS Population and Housing Census 2022, by zila")
    blob = fetch(WORKBOOK)
    log(f"    {len(blob):,} bytes from HDX")
    districts = read(blob)
    check(districts)

    records: list[dict[str, Any]] = []
    for row in districts:
        classified = sum(row["counts"].values())
        records.append(record(
            f"BGD-{row['name'].lower().replace(' ', '-')}",
            row["name"], level="admin2", parent="BGD",
            aliases=list(ALIASES.get(row["name"], ())),
            population=measure(row["population"], year=YEAR, source=SOURCE),
            religion=shares(row["counts"], total=classified) or gap(NOT_AVAILABLE),
            religion_year=YEAR, religion_note=NOTE,
            sources=[{"field": "population/religion", "name": SOURCE,
                      "url": URL, "license": LICENCE}]))

    out = args.out or PROCESSED / "bangladesh_district.json"
    write_json(out, records)
    log(f"  {len(records)} districts")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
