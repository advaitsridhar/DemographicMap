#!/usr/bin/env python3
"""Census tables the US Census Bureau standardises for countries that publish
theirs as documents.

The Bureau prepares "Subnational Population and Housing Census Tables" for
USAID's humanitarian bureau: a national census, extracted into one workbook per
country, with a schema that does not vary. Every sheet begins with the same
geography columns -- ``AREA_NAME``, ``ADM1_NAME``, ``ADM2_NAME``, ``ADM_LEVEL``
-- and then a block of value columns for one topic. So a single reader serves
every country in the series, which is why this is not two adapters:

* **Philippines**, 2020 Census of Population and Housing (Philippine Statistics
  Authority): religion and ethnicity for the country, 17 regions, and the
  provinces and highly urbanized cities beneath them.
* **Ethiopia**, 2007 Population and Housing Census (Central Statistical
  Agency): religion, ethnicity and language. 2007 because that is the last
  census Ethiopia completed -- the 2017 round was postponed and never held --
  so every record here carries that year rather than the year of the file.

Two things about the schema decide how this reads:

**The header is two rows deep.** Row 1 holds field names (``RLG_ORDX_B``) and
row 2 the aliases a person would recognise ("Orthodox, both sexes"). The names
are stable across countries and the aliases are what belongs on the map, so
both are read and the group label comes from the alias.

**Which columns are groups is a fact about the sheet, not a list to hardcode.**
The Philippine religion sheet names 128 individual denominations; Ethiopia's
names six. Hardcoding either would break on the other and, worse, would go
stale silently when the Bureau revises a workbook. So the group columns are
whatever is left after the geography columns and the denominator, and the run
prints how many it found. A column added upstream shows up in the log as a
count that moved, not as a figure that quietly changed.

**Sex is a check, not a dimension.** Ethiopia reports every religion three
times -- both sexes, females, males. Only the both-sexes column is a value
here; the other two are used to verify it, because a file whose sexes do not
add up is a file this reader has misunderstood.

Usage:
    python -m scripts.fetch_census.uscb --country PHL
    python -m scripts.fetch_census.uscb            # every country configured
"""

from __future__ import annotations

import argparse
import io
from collections import defaultdict
from dataclasses import dataclass, field as dc_field
from typing import Any

from ._shared import (
    NOT_AVAILABLE, PROCESSED, RAW, gap, http_get, http_json, log, measure,
    record, shares, write_json,
)

# HDX serves its API fully while blocking the HTML dataset pages behind a
# 2.3 kB stub, so the catalogue is reachable where the landing page is not.
HDX_API = "https://data.humdata.org/api/3/action"

# The HDX dataset holding a country's USCB tables is named in its config,
# because there is no deriving it. The obvious guess -- HDX's Common
# Operational Dataset for population, cod-ps-<iso3> -- is a different thing
# entirely: UNFPA and OCHA projections. It carries no USCB workbook for any of
# the 146 countries that have one, the Philippines and Ethiopia included. That
# was established by asking all 146, after assuming otherwise and breaking both
# working countries with the assumption.
#
# A dataset id is still worth resolving through rather than pinning a resource
# id. HDX gives the file a new resource id when the Bureau publishes a new
# extraction, and leaves the old one serving last year's figures until the day
# it 404s -- the failure mode being that nothing fails. The dataset it hangs
# on stays put.


def package(dataset: str) -> dict[str, Any]:
    body = http_json(f"{HDX_API}/package_show?id={dataset}")
    if not body.get("success"):
        raise SystemExit(f"HDX has no dataset {dataset!r}")
    return body["result"]


def dataset_url(dataset: str) -> str:
    """The HDX page a reader should be sent to for these tables."""
    return f"https://data.humdata.org/dataset/{dataset}"


def workbook_url(dataset: str) -> str:
    """The dataset's current USCB workbook.

    Raises rather than guessing: a dataset with no USCB workbook has none, and
    building a URL from the naming convention would turn that into a download
    error somewhere less obvious.
    """
    result = package(dataset)
    found = [r for r in result.get("resources", [])
             if "uscb" in (r.get("name") or "").lower()
             and (r.get("name") or "").lower().endswith(".xlsx")]
    if not found:
        names = ", ".join(sorted(r.get("name") or "?"
                                 for r in result.get("resources", []))[:8])
        raise SystemExit(f"{dataset} carries no USCB workbook. It holds: {names}")
    # Newest first, so a dataset carrying two extractions uses the later one.
    found.sort(key=lambda r: (r.get("name") or ""), reverse=True)
    if len(found) > 1:
        log(f"  {len(found)} USCB workbooks; using {found[0].get('name')!r}")
    return found[0]["url"]

# The columns every sheet in the series begins with, whatever its topic. Listed
# rather than detected: they are the schema, and a value column that happened
# to be named like one of them would be a silent loss.
GEOGRAPHY = {
    "AREA_NAME", "GEO_CONCAT", "GEO_MATCH", "CNTRY_NAME",
    "ADM1_NAME", "ADM2_NAME", "ADM3_NAME", "ADM4_NAME",
    "ADM_LEVEL", "USCBCMNT", "GENC_CODE", "FIPS_CODE",
    "NSO_CODE", "NSO_NAME",
}

# How far a group total may sit from the sheet's own denominator before it is
# worth naming, and how far before the file is refused outright.
#
# The two bounds do different jobs, and one bound could not do both. A reader
# that has misunderstood a sheet gets every area wrong; a census that suppressed
# a small group, or whose published total counts people its own categories do
# not, gets a handful wrong. So the test is the *shape* of the failure rather
# than its size alone: a few named areas are reported and kept, and either a
# large single gap or a widespread small one is refused.
#
# Measured on the Philippine ethnicity sheet, where MIMAROPA and its two
# Mindoro provinces fall 0.6% to 1.3% short -- and the region's shortfall is the
# sum of its provinces', which is what a real hole in the source looks like.
NOTABLE = 0.005
REFUSE_AREA = 0.05
REFUSE_SHARE = 0.10


@dataclass(frozen=True)
class Topic:
    """One sheet of a workbook, and the field it fills on this map.

    No column prefix: the first version of this carried one per topic and was
    wrong on its second file, because Ethiopia's ethnic-group columns are not
    named "ETH_". Which columns hold groups is a fact the sheet states -- they
    are the ones that are not geography -- and asking it is both simpler and
    the thing this module's docstring already said to do.

    Year, source and note override the country's where a topic does not come
    from the same place as the rest of the file. Burma is why. Its Age-Sex
    sheet is the 2014 Population and Housing Census, and its Ethnicity sheet
    is the Department of Population's 2018 Township Profiles with a 2017
    reference date -- because the 2014 census's ethnicity tables were never
    published. Taking the country's year for both would date 2017 figures to
    2014 and attribute them to a census that did not release them.
    """
    sheet: str
    field: str
    # The column holding everyone the question was asked of, named outright
    # where denominator() cannot find it. That function looks for "population"
    # in the alias, which is right for every census here and wrong for the DRC:
    # its universe is TRB_SSIZE / "Sample size", the number of heads of
    # household the survey reached. Left unnamed it would be read as a tribe --
    # the largest one in the country, since it is the sum of all the others.
    denominator: str = ""
    # Set only where one sheet holds two topics. Burma's is called "Ethnicity"
    # and carries the religion columns as well, so "everything that is not
    # geography" collected both and the shares came to 3.05 times the
    # population. Where a sheet holds one topic this stays empty and the
    # columns are still whatever is left after the geography -- which is what
    # keeps Ethiopia working, whose ethnic-group columns are not named "ETH_".
    prefix: str = ""
    year: int | None = None
    source: str = ""
    note: str = ""


@dataclass(frozen=True)
class Country:
    iso3: str
    name: str
    year: int
    source: str
    licence: str
    out: str
    # Which ADM_LEVEL becomes which layer of this map. The Bureau's levels are
    # the country's own, so this is a fact per country: the Philippines' first
    # order is its 17 regions and its second the provinces, while Ethiopia's
    # first order is its regions and its second the zones.
    levels: dict[int, str]
    topics: tuple[Topic, ...]
    note: str
    # The HDX dataset holding this country's tables. Both the workbook and the
    # citation link come from it, so there is one identifier per country rather
    # than a resource UUID and a hand-written slug that can disagree -- and the
    # slugs they replace ("philippines-subnational-population-statistics") were
    # never fetched, so nothing had ever checked they resolve.
    dataset: str = ""
    # How far this country's areas may fall short of their own published
    # totals before the file is refused. Defaults are the module's; raising
    # them is a claim about the source and needs the evidence written down,
    # because the whole point of the bounds is that a reader which has
    # misunderstood a sheet gets *every* area wrong.
    refuse_area: float = REFUSE_AREA
    refuse_share: float = REFUSE_SHARE
    # An ADM_LEVEL whose rows the source lists but leaves blank, to be built by
    # adding up the level below. Ukraine is the case: its 2001 language table
    # is published by rayon, and the 25 oblast rows above them are empty, so
    # the country's first order carried nothing at all. The file's own
    # ADM1_NAME column says which oblast each rayon belongs to, so the addition
    # is the source's arithmetic rather than this map's invention.
    #
    # Only areas with no figures of their own are built this way. Kyiv and
    # Sevastopol publish theirs directly and keep them.
    sum_into: int | None = None
    aliases: dict[str, tuple[str, ...]] = dc_field(default_factory=dict)
    # Census areas geoBoundaries draws no boundary of their own for, as
    # (parent, area) pairs -- a name alone is not an address, and the
    # Philippines has a Quezon that is a province and a Quezon that is a city
    # 130 km away.
    #
    # This is not a list of everything unmatched. It is the subset the boundary
    # file folds away *by design* -- a highly urbanized city drawn inside the
    # province around it, Addis Ababa's ten sub-cities drawn as one "Region 14"
    # -- where saying so is what stops the row from reaching a shape that is
    # not its own. Left undeclared, "Cebu City" and "Province Of Cebu" both
    # normalise to "cebu", both match outright, and the collision resolver
    # treats two outright matches as one place listed twice: the last one wins
    # and a city of 964,000 quietly wears a province's figures, or the reverse.
    # Three of those were happening here (Cebu, Iloilo, Quezon City).
    #
    # An area the boundary file merely lacks is deliberately *not* listed --
    # Sultan Kudarat is a province with no CGAZ shape at all, Sidama a region
    # created after CGAZ's Ethiopian vintage. Those are omissions that an
    # upstream release could fix, and a declaration here would then be a lie
    # that suppresses a match. A declaration should only ever be doing work.
    no_shape: frozenset[tuple[str, str]] = frozenset()
    # Whether this file's figures count people. They do everywhere but the
    # DRC, whose survey counts the heads of household it sampled -- 31,755 of
    # them, for 119 million people. Publishing that beside Pakistan's census
    # counts would put a sample and a population in one field with nothing to
    # tell them apart, so where this is false only the share is published.
    # The share is what the survey measures; the count is how many doors it
    # knocked on. Shares without counts are already a shape this map has: the
    # whole-country rows from the Factbook carry exactly that.
    counts_are_people: bool = True


HDX = "https://data.humdata.org/dataset"

PHILIPPINES = Country(
    iso3="PHL",
    name="Philippines",
    year=2020,
    source=("Philippine Statistics Authority, 2020 Census of Population and "
            "Housing, prepared as subnational tables by the U.S. Census Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="809dfb22-77f4-482c-8560-79b07d20fc15",
    out="philippines_province.json",
    levels={1: "admin1", 2: "admin2"},
    topics=(Topic("Religion", "religion"),
            Topic("Ethnicity", "ethnicity")),
    # What geoBoundaries calls the same area. The census writes a region's
    # full name and the boundary file its initials, which no matching rule
    # should bridge on its own -- "NCR" and "National Capital Region" share no
    # word, and a rule loose enough to join them would join a great deal else.
    #
    # The cities are not here but in no_shape below: they are absent rather
    # than misnamed, and an alias cannot conjure a boundary.
    aliases={
        "Bangsamoro Autonomous Region Of Muslim Mindanao": ("ARMM",),
        "Cordillera Administrative Region": ("CAR",),
        "National Capital Region": ("NCR",),
        "Caraga Region": ("Caraga",),
        "Province Of Cebu": ("Cebu",),
        "Province Of Cotabato": ("Cotabato",),
        "Province Of Iloilo": ("Iloilo",),
        "Mountain": ("Mountain Province",),
        "Davao De Oro": ("Compostela Valley",),
        "Manila": ("City of Manila",),
    },
    # geoBoundaries draws Metro Manila as four numbered districts, so fifteen
    # of its sixteen cities have no shape (the City of Manila is the exception
    # and is aliased above), and it folds every other highly urbanized city
    # into the province around it. Three of those cities share a normalised
    # name with a province and were silently taking its shape.
    no_shape=frozenset((
        ("National Capital Region", "Caloocan"),
        ("National Capital Region", "Las Piñas"),
        ("National Capital Region", "Makati"),
        ("National Capital Region", "Malabon"),
        ("National Capital Region", "Mandaluyong"),
        ("National Capital Region", "Marikina"),
        ("National Capital Region", "Muntinlupa"),
        ("National Capital Region", "Navotas"),
        ("National Capital Region", "Parañaque"),
        ("National Capital Region", "Pasay"),
        ("National Capital Region", "Pasig"),
        ("National Capital Region", "Pateros"),
        ("National Capital Region", "Quezon"),
        ("National Capital Region", "San Juan"),
        ("National Capital Region", "Taguig"),
        ("National Capital Region", "Valenzuela"),
        ("Central Luzon", "Angeles"),
        ("Central Luzon", "Olongapo"),
        ("Calabarzon", "Lucena"),
        ("Mimaropa", "Puerto Princesa"),
        ("Western Visayas", "Bacolod"),
        ("Western Visayas", "Iloilo City"),
        ("Central Visayas", "Cebu City"),
        ("Central Visayas", "Lapu-Lapu"),
        ("Central Visayas", "Mandaue"),
        ("Eastern Visayas", "Tacloban"),
        ("Cordillera Administrative Region", "Baguio"),
        ("Caraga Region", "Butuan"),
        ("Northern Mindanao", "Cagayan De Oro"),
        ("Northern Mindanao", "Iligan"),
        ("Soccsksargen", "General Santos"),
        ("Davao Region", "Davao"),
        ("Zamboanga Peninsula", "Zamboanga"),
    )),
    note=("2020 Census of Population and Housing. The census records religious "
          "affiliation as the individual church or denomination a person names, "
          "so the groups here are as published rather than collapsed into "
          "traditions -- the alternative would decide, on this map's authority "
          "rather than the census's, which churches count as one religion."),
)

ETHIOPIA = Country(
    iso3="ETH",
    name="Ethiopia",
    year=2007,
    source=("Central Statistical Agency of Ethiopia, 2007 Population and "
            "Housing Census, prepared as subnational tables by the U.S. Census "
            "Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="5438946a-51b7-44d6-9b76-aeafdb4dc4d3",
    out="ethiopia_region.json",
    levels={1: "admin1", 2: "admin2"},
    topics=(Topic("Religion", "religion"),
            Topic("Ethnicity", "ethnicity"),
            Topic("Language", "language")),
    # The Bureau transliterates from Amharic and geoBoundaries does not, so
    # most of Ethiopia's regions reach the map only by declaration: "Oromīya"
    # and "Oromia" differ by a letter, "Sumalē" and "Somali" by three, and
    # "Yedebub Bihēroch Bihēreseboch Na Hizboch" is the region the boundary
    # file simply calls SNNPR. Folding the diacritics is not enough for any of
    # them, and a rule loose enough to bridge the last would bridge anything.
    #
    # Not listed: Addis Ababa's ten sub-cities, which geoBoundaries draws as
    # the single shape "Region 14", and the special weredas it draws as one
    # "Special Woreda". Those are absent shapes, not wrong names.
    aliases={
        "Ādīs Ābeba": ("Addis Ababa",),
        "Āfar": ("Afar",),
        "Āmara": ("Amhara",),
        "Bīnshangul Gumuz": ("Beneshangul Gumu", "Benishangul Gumuz"),
        "Dirē Dawa": ("Dire Dawa",),
        "Gambēla Hizboch": ("Gambela",),
        "Hārerī Hizb": ("Hareri",),
        "Oromīya": ("Oromia",),
        "Sumalē": ("Somali",),
        "Yedebub Bihēroch Bihēreseboch Na Hizboch": ("SNNPR",),
        # Zones, where the two spell the same place differently.
        "Kellem Wellega": ("Kelem Wellega",),
        "Kembata Tembaro": ("KT",),
        "Mezhenger": ("Majang",),
        "Basketo Special Wereda": ("Basketo",),
        "Harari": ("Hareri",),
        "Oromiya": ("Oromia",),
        "Southwest Shuwa": ("South West Shewa",),
        "Welayita": ("Wolayita",),
    },
    # Addis Ababa's ten sub-cities are one shape called "Region 14", and the
    # special weredas and town administrations are one called "Special Woreda".
    # Basketo is not here: it has a shape of its own and is aliased above.
    # "Jimma Town Special Wereda" is the one that was doing damage -- it
    # reached the zone of Jimma by prefix, 350 km of countryside wearing a
    # town's figures, and was refused only because the zone's own row got
    # there first.
    no_shape=frozenset((
        ("Ādīs Ābeba", "Addis Ketema"),
        ("Ādīs Ābeba", "Akaki Kaliti"),
        ("Ādīs Ābeba", "Arada"),
        ("Ādīs Ābeba", "Bole"),
        ("Ādīs Ābeba", "Gulele"),
        ("Ādīs Ābeba", "Kirkos"),
        ("Ādīs Ābeba", "Kolfe Keranyo"),
        ("Ādīs Ābeba", "Lideta"),
        ("Ādīs Ābeba", "Nefas Silk Lafto"),
        ("Ādīs Ābeba", "Yeka"),
        ("Bīnshangul Gumuz", "Mao Komo Special Wereda"),
        ("Bīnshangul Gumuz", "Pawe Special Wereda"),
        ("Gambēla Hizboch", "Etang Special Wereda"),
        ("Oromīya", "Adama Special Wereda"),
        ("Oromīya", "Burayu Special Wereda"),
        ("Oromīya", "Jimma Town Special Wereda"),
        ("Tigray", "Mekele Town Special Wereda"),
        ("Āmara", "Argoba Special Wereda"),
        ("Āmara", "Bahir Dar Special Wereda"),
        ("Yedebub Bihēroch Bihēreseboch Na Hizboch", "Hawassa City Administration"),
    )),
    note=("2007 Population and Housing Census -- the last census Ethiopia has "
          "completed. The 2017 round was postponed and never held, so this is "
          "the most recent measurement in existence, not the most recent "
          "available to this project. Read it as a description of 2007."),
)

MYANMAR = Country(
    iso3="MMR",
    name="Myanmar",
    year=2014,
    source=("Department of Population, Ministry of Labour, Immigration and "
            "Population, 2014 Myanmar Population and Housing Census, prepared "
            "as subnational tables by the U.S. Census Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="burma-subnational-boundaries-and-tabular-data",
    out="myanmar_state.json",
    levels={1: "admin1", 2: "admin2"},
    # One sheet, two topics. It is called "Ethnicity" and carries the religion
    # columns beside the ethnic ones, so reading "everything that is not
    # geography" collected both and the shares came to 3.05 times the
    # population -- 47.8 million of ethnicity plus 49.0 million of religion
    # plus the religion denominator, all summed as though they were one
    # question. The prefixes are what separate them here.
    topics=(
        # Not the census. The 2014 round asked about ethnicity and its results
        # were never published -- the tables were withheld -- so the Bureau
        # uses the Department of Population's township profiles instead, and
        # the file says so in its data dictionary while the sheet is still
        # called "Ethnicity". Dating these to 2014 would attribute them to a
        # census that refused to release them, which is the notable fact about
        # ethnicity data in Myanmar and the last thing to bury.
        Topic("Ethnicity", "ethnicity", prefix="ETH_", year=2017,
              source=("Department of Population, 2018 Township Profiles, "
                      "Table 14: Ethnic Nationalities Living, prepared as "
                      "subnational tables by the U.S. Census Bureau"),
              note=("2018 Township Profiles, with a reference date of 1 April "
                    "2017. The 2014 census collected ethnicity and its results "
                    "were never released, so these are the Department of "
                    "Population's own later figures rather than census "
                    "counts. Myanmar recognises 135 official ethnic groups "
                    "and this table names 40; the Rohingya are not among "
                    "either, having been excluded from enumeration as an "
                    "ethnicity in 2014.")),
        # Religion, on the other hand, the 2014 census did publish, and these
        # columns carry its own denominator of 49.0 million against the
        # ethnicity table's 47.8 million -- two different populations in one
        # sheet, which is the other reason they cannot be read together.
        Topic("Ethnicity", "religion", prefix="RLG_"),
    ),
    # None needed for the states and regions: the census writes "KACHIN STATE"
    # where geoBoundaries writes "Kachin", and norm() drops the word "state"
    # on both sides. Two boundary names are misspelled rather than differently
    # spelled, and are corrected in common.MISSPELLED so the label a reader
    # sees is fixed along with the join.
    #
    # The districts are another matter: both sides romanise from Burmese and
    # neither is wrong, so these are declared pair by pair against the two
    # lists rather than bridged by a rule. Not listed, because they are absent
    # rather than differently spelled: the six self-administered zones and
    # divisions, which geoBoundaries draws as ordinary townships, and Nay Pyi
    # Taw, which it has no first-order shape for at all.
    aliases={
        "Bawlakhe": ("Bawlake",),
        "Haka": ("Hakha",),
        "Kawthaung": ("Kawthoung",),
        "Kyaunkpyu": ("Kyaukpyu",),
        "Langhko": ("Langkho",),
        "Loilem": ("Loilen",),
        "Myauk U": ("Mrauk-U",),
        "Pharpon": ("Hpapun",),
        "Tharrawaddy": ("Thayarwady",),
        "Yinmarpin": ("Yinmarbin",),
    },
    note=("2014 Myanmar Population and Housing Census unless a field says "
          "otherwise."),
)

UKRAINE = Country(
    iso3="UKR",
    name="Ukraine",
    year=2001,
    source=("State Statistics Service of Ukraine, All-Ukrainian Population "
            "Census 2001, prepared as subnational tables by the U.S. Census "
            "Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="ukraine-subnational-boundaries-and-tabular-data",
    out="ukraine_oblast.json",
    # Oblasts only, though the figures live in the rayons and are summed out of
    # them. Publishing the 661 rayons put 605 rows nowhere and 56 onto shapes
    # by prefix and containment against an adjectival form -- matches no one
    # can check one by one, which is the kind this map refuses on principle
    # even when most of them are probably right. The oblasts match one for one
    # and reconcile to the census's own national total exactly.
    levels={1: "admin1"},
    # The Language sheet, not Nationality-Language. That second sheet is the
    # cross-tabulation of the two -- 1,619 columns, every nationality against
    # every native language -- which is a different and much larger claim than
    # this map has a field for. The flat sheet is the one that answers "what
    # is spoken here".
    topics=(Topic("Language", "language"),),
    # Only the share bound moves, and only because 27 areas is a coarse
    # denominator: one oblast is 3.7% of the count, so three of them tripping
    # a 10% test says almost nothing about whether the sheet was understood.
    # The size bound stays at the module's default and is not close to being
    # tested -- the worst oblast is 0.7% short, where 5% is the refusal.
    #
    # An earlier version widened the size bound to 8% as well, on evidence
    # from the rayons: 105 of 663 fell short by up to 6.9%. Those rayons are
    # no longer published, so that evidence no longer describes anything this
    # adapter emits, and the bound goes back.
    refuse_share=0.20,
    # The oblast rows are blank and the rayons beneath them are not, so the
    # first order is built by addition. That is what makes Ukraine joinable at
    # all: its 27 oblasts match geoBoundaries one for one once aliased, while
    # its 661 rayons would each need a name the boundary file spells
    # differently.
    sum_into=1,
    # The Bureau romanises from Ukrainian and geoBoundaries uses the English
    # exonyms, so every one of the 27 needs declaring: "CHERKAS'KA OBLAST'"
    # and "Cherkasy Oblast" share no word that norm() leaves standing.
    aliases={
        "Avtonomna Respublika Krym": ("Autonomous Republic of Crimea",),
        "Misto Sevastopol’": ("Sevastopol",),
        "Misto Kyyiv": ("Kyiv",),
        "Cherkas’Ka Oblast’": ("Cherkasy Oblast",),
        "Chernihivs’Ka Oblast’": ("Chernihiv Oblast",),
        "Chernivets’Ka Oblast’": ("Chernivtsi Oblast",),
        "Dnipropetrovs’Ka Oblast’": ("Dnipropetrovsk Oblast",),
        "Donets’Ka Oblast’": ("Donetsk Oblast",),
        "Ivano-Frankivs’Ka Oblast’": ("Ivano-Frankivsk Oblast",),
        "Kharkivs’Ka Oblast’": ("Kharkiv Oblast",),
        "Khersons’Ka Oblast’": ("Kherson Oblast",),
        "Khmel’Nyts’Ka Oblast’": ("Khmelnytskyi Oblast",),
        "Kirovohrads’Ka Oblast’": ("Kirovohrad Oblast",),
        "Kyyivs’Ka Oblast’": ("Kyiv Oblast",),
        "Luhans’Ka Oblast’": ("Luhansk Oblast",),
        "L’Vivs’Ka Oblast’": ("Lviv Oblast",),
        "Mykolayivs’Ka Oblast’": ("Mykolaiv Oblast",),
        "Odes’Ka Oblast’": ("Odessa Oblast",),
        "Poltavs’Ka Oblast’": ("Poltava Oblast",),
        "Rivnens’Ka Oblast’": ("Rivne Oblast",),
        "Sums’Ka Oblast’": ("Sumy Oblast",),
        "Ternopil’S’Ka Oblast’": ("Ternopil Oblast",),
        "Vinnyts’Ka Oblast’": ("Vinnytsia Oblast",),
        "Volyns’Ka Oblast’": ("Volyn Oblast",),
        "Zakarpats’Ka Oblast’": ("Zakarpattia Oblast",),
        "Zaporiz’Ka Oblast’": ("Zaporizhia Oblast",),
        "Zhytomyrs’Ka Oblast’": ("Zhytomyr Oblast",),
    },
    note=("All-Ukrainian Population Census 2001 -- the only census independent "
          "Ukraine has held. The question is native language, which the census "
          "asked separately from nationality; the two differ substantially and "
          "this is the language answer. Read it as a description of 2001."),
)

# Ukraine was configured and held back for one release: its figures reconciled
# and 58 of 663 areas joined, because the source publishes at rayon level and
# the census romanises Ukrainian adjectivally where geoBoundaries uses the
# short exonym -- "BAKHCHYSARAYS'KYY RAYON" against "Bakhchysarai", 661 times.
# Summing each oblast's own rayons is what got it in, and the oblasts match one
# for one.
PAKISTAN = Country(
    iso3="PAK",
    name="Pakistan",
    year=2017,
    source=("Pakistan Bureau of Statistics, 2017 Population and Housing "
            "Census, Table 11: Population by mother tongue, sex, and "
            "rural/urban, prepared as subnational tables by the U.S. Census "
            "Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="pakistan-subnational-population-and-housing-data-tables",
    out="pakistan_language.json",
    # Level 2 is the 36 divisions and level 3 the 155 districts. It is the
    # districts geoBoundaries draws as Pakistan's second order, so reading
    # level 2 here would put a division's figures on a district's shape --
    # every one of them wrong, and none of them visibly so.
    levels={1: "admin1", 3: "admin2"},
    # Religion is deliberately not read, though this workbook carries it.
    # scripts/fetch_census/pakistan.py already publishes it from the same
    # census table, and two files claiming one shape with figures that
    # disagree by a single person would send both to a gap under the
    # agreement rule. Language is a field nothing else fills, so it is added
    # where religion could only be risked.
    topics=(Topic("Mother Tongue", "language", prefix="LNG_"),),
    # The word "district" is dropped by norm(), so 105 of 126 districts
    # reached their shape with no help at all. These nine are the ones the
    # two sources spell differently, and each was read off the pair of
    # leftovers rather than guessed: after the first join, exactly these
    # census rows had reached no shape and exactly these shapes had no row.
    #
    # Six are transliteration (Killa/Qilla, Kambar/Qambar, Jaffarabad with
    # two f's, Batagram with one t, Sheikhupura and Vehari with the vowels
    # moved). Shaheed Benazirabad is different in kind: the district was
    # renamed from Nawabshah in 2008 and geoBoundaries still carries the old
    # name, so this alias is a date rather than a spelling.
    aliases={
        "Batagram District": ("Battagram",),
        "Jaffarabad District": ("Jafarabad",),
        "Kambar Shahdadkot District": ("Qambar Shahdadkot",),
        "Killa Abdullah District": ("Qilla Abdullah",),
        "Killa Saifullah District": ("Qilla Saifullah",),
        "Naushahro Feroze District": ("Naushehro Feroze",),
        "Shaheed Benazirabad District": ("Nawabshah",),
        "Sheikhupura District": ("Sheikhpura",),
        "Vehari District": ("Vihari",),
    },
    # The census counted the Federally Administered Tribal Areas as a
    # first-order area; they were merged into Khyber Pakhtunkhwa in 2018 and
    # geoBoundaries draws the seven that remain. Declared, so it reads as a
    # boundary that no longer exists rather than as a row this map lost.
    no_shape=frozenset(
        (("Pakistan", "Federally Administered Tribal Areas"),)),
    note=("2017 Population and Housing Census. The question is mother tongue, "
          "which is the language of the household a person grew up in rather "
          "than the language they speak now, and Pakistan's nine named "
          "tongues leave a tenth column of Other -- 2.3% nationally, and the "
          "place where Shina, Balti and Khowar are counted without being "
          "named."),
)


CENTRAL_AFRICAN_REPUBLIC = Country(
    iso3="CAF",
    name="Central African Republic",
    year=2003,
    source=("Bureau Central du Recensement, Troisième Recensement Général de "
            "la Population et de l'Habitation (RGPH03), prepared as "
            "subnational tables by the U.S. Census Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="car-subnational-boundaries-and-tabular-data",
    out="car_prefecture.json",
    # The straightforward case, and worth saying so after Pakistan: the file's
    # 17 first-order areas and 72 second-order areas are exactly what
    # geoBoundaries draws, so the levels are the levels.
    levels={1: "admin1", 2: "admin2"},
    # Three sheets, one question each, and no prefix. Each carries its own
    # ETH_/RLG_/LNG_ columns, but naming a prefix here would silently drop any
    # column that did not match it, and the evidence that nothing is being
    # wrongly included or excluded is stronger without one: all three sheets
    # reconcile to within a single person of their own published totals.
    topics=(Topic("Ethnicity", "ethnicity"),
            Topic("Religion", "religion"),
            Topic("Language", "language")),
    # 17 of 17 prefectures and 68 of 72 sub-prefectures reach a shape on their
    # own name. These four are the whole of the disagreement, read off the two
    # lists of leftovers -- four rows with no shape against four shapes with no
    # row, pairing one to one. Two are a doubled or dropped letter, one an
    # accent the census does not write, and Ndjoukou loses its initial N in the
    # boundary file.
    aliases={
        "Aba": ("Abba",),
        "Bossemptele": ("Bossemtélé",),
        "Nanga-Boguila": ("Nagha Boguila",),
        "Ndjoukou": ("Djoukou",),
    },
    note=("2003 census (RGPH03). The ethnicity table is the census's own "
          "\"Ethnie-Nationalité\", which codes residents who are not "
          "Central African by nationality rather than by ethnic group: "
          "Cameroonian, Chadian, French and Lebanese sit in the same column "
          "set as Gbaya and Banda. Those categories are 1.8% of the "
          "population between them and the nine local groups are 96.1%, so "
          "this is an ethnicity table with a foreign tail rather than the "
          "nationality table Syria's sheet turned out to be -- and dropping "
          "the tail would stop the composition summing to the population it "
          "is a composition of. One local category, Hausa, is written "
          "\"Haoussa/Musulman\" in the original, defining an ethnic group "
          "partly by religion; it is published as the census wrote it."),
)


MALI = Country(
    iso3="MLI",
    name="Mali",
    year=2009,
    source=("Institut National de la Statistique (INSTAT), 4ème Recensement "
            "Général de la Population et de l'Habitat 2009, Tableau S-3: "
            "Population Résidente de 6 Ans et Plus Selon la Langue, prepared "
            "as subnational tables by the U.S. Census Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="mali-subnational-population-and-housing-data-tables-with-"
            "administrative-boundaries",
    out="mali_region.json",
    # Regions only. The file lists all 50 cercles and leaves every one of
    # them blank, so claiming the second order here would be claiming a level
    # that carries nothing -- the reverse of Ukraine, whose figures were at
    # the second order and whose first was the empty one. geoBoundaries draws
    # both, and the cercles keep a visible gap.
    levels={1: "admin1"},
    topics=(Topic("Language", "language"),),
    aliases={},
    note=("2009 census. The shares are of the resident population aged 6 and "
          "over, not of everybody: the census table is titled \"Population "
          "Résidente de 6 Ans et Plus Selon la Langue\" and its columns sum "
          "to 11,109,312 against a counted population of about 14.5 million. "
          "The missing quarter is children under six, who were not asked. "
          "That gap was the reason this file was held back a release -- a "
          "quarter of a country unaccounted for is either a stated universe "
          "or a misread sheet, and which one it was is a question only the "
          "source's own citation could answer."),
)


DEMOCRATIC_REPUBLIC_OF_THE_CONGO = Country(
    iso3="COD",
    name="Democratic Republic of the Congo",
    year=2012,
    source=("Institut National de la Statistique, Enquête 1-2-3 (2005 and "
            "2012), prepared as subnational tables by the U.S. Census Bureau"),
    licence="CC BY-IGO, published via HDX",
    dataset="democratic-republic-of-the-congo-subnational-boundaries-and-"
            "tabular-data",
    out="drc_province.json",
    # Provinces only, though the file carries 164 districts. 31,755 households
    # spread over 26 provinces is about 1,200 each, which supports a
    # provincial estimate; spread over the districts it is about 194 each,
    # which does not support anything. The second order is read and not
    # published, and that is a decision about sample size rather than about
    # names -- every district would have joined.
    levels={1: "admin1"},
    # One sheet, two questions, as Burma's was -- so each topic takes its own
    # prefix, and both take the same denominator, which belongs to neither.
    topics=(Topic("Tribe and Religion", "ethnicity", prefix="TRB_",
                  denominator="TRB_SSIZE"),
            Topic("Tribe and Religion", "religion", prefix="RLG_",
                  denominator="TRB_SSIZE")),
    # The figures are households, not people. See counts_are_people.
    counts_are_people=False,
    aliases={},
    note=("Enquête 1-2-3, pooling the 2005 and 2012 rounds. This is the only "
          "source on this map that is a survey rather than a census, and it "
          "is reported differently for that reason: the shares are of the "
          "31,755 heads of household the survey reached, about 1,200 per "
          "province, and no count is published because the count is of "
          "households sampled rather than of people. A provincial share here "
          "carries sampling error that a census count does not, and it "
          "describes the household head rather than the household. Two "
          "survey rounds seven years apart are pooled into one figure by the "
          "source, so it dates to neither year exactly."),
)


COUNTRIES: dict[str, Country] = {
    c.iso3: c for c in (PHILIPPINES, ETHIOPIA, MYANMAR, UKRAINE,
                        PAKISTAN, CENTRAL_AFRICAN_REPUBLIC, MALI,
                        DEMOCRATIC_REPUBLIC_OF_THE_CONGO)}


def discover(limit: int, sheets: bool) -> int:
    """Every USCB workbook on HDX, found by asking for the resources directly.

    Searching datasets cannot answer this. The workbook hangs off a dataset
    whose title says nothing about the Bureau, and the first attempt at this --
    deriving the dataset name from the ISO3 code -- reported that none of the
    146 countries in HDX's population collection had one, which was true and
    useless, because it was looking in the wrong collection entirely.

    So this asks CKAN for resources by name and works back to the dataset each
    belongs to. What then decides whether a country is worth a config is its
    sheet names, and those are only knowable by opening the file: Indonesia has
    a workbook and it holds no religion and no ethnicity, only a four-bucket
    first-language split.
    """
    import openpyxl

    body = http_json(f"{HDX_API}/resource_search?query=name:uscb&limit=1000")
    rows = body.get("result", {}).get("results", [])
    books = [r for r in rows if (r.get("name") or "").lower().endswith(".xlsx")]
    seen: dict[str, dict[str, Any]] = {}
    for row in books:
        seen.setdefault(row.get("package_id") or row.get("id"), row)
    log(f"{len(rows)} resources named for the Bureau, {len(books)} of them "
        f"workbooks, across {len(seen)} datasets")

    order = sorted(seen.items(), key=lambda kv: kv[1].get("name") or "")
    for i, (pkg, row) in enumerate(order):
        if limit and i >= limit:
            log(f"  ...stopping at {limit} of {len(order)}; --limit 0 for all")
            break
        try:
            meta = package(pkg)
        except SystemExit as why:
            log(f"  {row.get('name')}: {why}")
            continue
        line = f"  {row.get('name')}  {meta.get('name')}  ({meta.get('title')})"
        if not sheets:
            log(line)
            continue
        blob = http_get(row["url"], binary=True, cache_dir=RAW / "uscb")
        book = openpyxl.load_workbook(io.BytesIO(blob), read_only=True,
                                      data_only=True)
        names = list(book.sheetnames)
        book.close()
        # Named in full rather than filtered to the three topics this map
        # wants: a topic nobody thought to look for is exactly what a listing
        # surfaces.
        log(f"{line}\n      {len(blob):,} bytes, sheets: {', '.join(names)}")
    return 0


def inspect(dataset: str, wanted: list[str]) -> int:
    """Everything a config needs, read off the workbook rather than guessed.

    Writing a Country by hand needs four facts the sheet names do not carry:
    which ADM_LEVEL is this map's admin1 and which its admin2, what the group
    labels actually say, what census year the tables are from, and what the
    areas are called so they can be aliased to the boundary file. All four are
    in the file. None is in the catalogue.
    """
    import openpyxl

    blob = http_get(workbook_url(dataset), binary=True, cache_dir=RAW / "uscb")
    book = openpyxl.load_workbook(io.BytesIO(blob), read_only=True,
                                  data_only=True)
    log(f"{dataset}: {len(blob):,} bytes, {len(book.sheetnames)} sheets")

    # The metadata sheet is where the census year lives, in table identifiers
    # like ET_RELIGION_2007census. The filename's date is an extraction date
    # and has been mistaken for the census year before.
    #
    # The data dictionary is more important still. It gives every column a
    # definition and a source, and that is what says whether a sheet holds
    # what its name suggests -- Syria's "Ethnicity" turns out to list Syrian,
    # Palestinian, European and American, which is nationality. Publishing
    # that as ethnicity would tell a reader Syria is ethnically uniform, which
    # is not what the census measured or said.
    dictionary: dict[str, list[str]] = {}
    for name in book.sheetnames:
        if name.strip().lower() not in ("metadata", "data dictionary"):
            continue
        rows = [list(r) for r in book[name].iter_rows(values_only=True)]
        log(f"\n--- {name}: {len(rows)} rows ---")
        for row in rows[:20]:
            cells = [str(v).strip() for v in row if v is not None]
            if cells:
                log("  " + " | ".join(c[:70] for c in cells)[:220])
        if name.strip().lower() == "data dictionary":
            for row in rows:
                cells = [("" if v is None else str(v).strip()) for v in row]
                if cells and cells[0]:
                    dictionary.setdefault(cells[0], cells[1:])

    for name in book.sheetnames:
        if name.strip().lower() in ("metadata", "data dictionary"):
            continue
        if wanted and not any(w.lower() in name.lower() for w in wanted):
            continue
        rows = [list(r) for r in book[name].iter_rows(values_only=True)]
        if len(rows) < 3:
            log(f"\n--- {name}: {len(rows)} rows, nothing under the header ---")
            continue
        names, aliases = columns(rows)
        if "ADM_LEVEL" not in names or "AREA_NAME" not in names:
            log(f"\n--- {name}: not a geography sheet "
                f"(no ADM_LEVEL/AREA_NAME) ---")
            continue
        by_level: dict[Any, list[tuple[str, str]]] = {}
        for row in rows[2:]:
            level, area_name, parent = area(row, names)
            by_level.setdefault(level, []).append((parent, area_name))
        log(f"\n--- {name}: {len(rows) - 2} rows, sexed={sexed(names)} ---")
        for level in sorted(by_level, key=lambda v: (v is None, v)):
            here = by_level[level]
            # Every name where there are few enough to read, because writing
            # the aliases is the actual job and a sample of eight cannot do
            # it: the census says "CHERKAS'KA OBLAST'" where the boundary file
            # says "Cherkasy Oblast", and each of the 27 needs declaring.
            shown = ", ".join(n for _p, n in here) if len(here) <= 40 \
                else ", ".join(n for _p, n in here[:8]) + ", ..."
            log(f"  ADM_LEVEL {level}: {len(here)} areas  {shown}")
        total = denominator(names, aliases)
        found = groups(names, aliases, total, sexed(names))
        log(f"  denominator: {aliases[total]!r}" if total is not None
            else "  denominator: none found")
        log(f"  {len(found)} group columns:")
        # groups() returns {column index: label}, and the label is what the
        # map publishes -- for a sexed sheet it is the alias with ", both
        # sexes" already stripped, so printing the raw alias would show
        # something the adapter never emits.
        for index, label in list(found.items())[:24]:
            log(f"    {names[index]}  ->  {label!r}")
        if len(found) > 24:
            log(f"    ...and {len(found) - 24} more")
        # What the source says these columns are. A label is a word; the
        # dictionary row is the claim the map would be repeating.
        for index in list(found)[:3]:
            entry = dictionary.get(names[index])
            if entry:
                # 400 rather than 150: the third cell is the source
                # citation, and Mali's was cut at "Tableau S-3: Population
                # R..." -- exactly the clause naming which population the
                # table counts, which is the question a truncated citation
                # cannot answer and the whole reason for printing it.
                log(f"    {names[index]} says: "
                    + " | ".join(c[:400] for c in entry if c))
        # The whole country's row, every column, against the total it
        # publishes. This is what says whether the columns partition a
        # population or overlap: Burma's forty sum to 3.05 times its own
        # published total, and the only way to see why is to read all forty
        # with their shares beside them.
        def breakdown(row: list[Any], why: str) -> None:
            level, area_name, _parent = area(row, names)
            published = number(row[total]) if total is not None else None
            summed = sum(v for v in (number(row[i]) for i in found)
                         if v is not None)
            ratio = (summed / published) if published else 0.0
            log(f"\n  {why}: {area_name} (level {level}) "
                f"published={published:,.0f} summed={summed:,.0f} "
                f"ratio={ratio:.3f}" if published else
                f"\n  {why}: {area_name} (level {level}) no published total")
            for index, label in found.items():
                value = number(row[index])
                if value is None:
                    continue
                share = (100.0 * value / published) if published else 0.0
                log(f"    {names[index]:<14} {label[:34]:<34} "
                    f"{value:>13,.0f}  {share:6.2f}%")

        national = next((r for r in rows[2:] if area(r, names)[0] == 0), None)
        if national is not None:
            breakdown(national, "whole country")
        # And the areas furthest from their own total, which is where a
        # misread sheet shows itself first.
        def gap_of(row: list[Any]) -> float:
            published = number(row[total]) if total is not None else None
            if not published:
                return 0.0
            summed = sum(v for v in (number(row[i]) for i in found)
                         if v is not None)
            return abs(summed - published) / published
        worst = sorted((r for r in rows[2:] if area(r, names)[0]),
                       key=gap_of, reverse=True)[:2]
        for row in worst:
            if gap_of(row) > 0.001:
                breakdown(row, "furthest from its own total")
    book.close()
    return 0


def sheet_rows(book, name: str) -> list[list[Any]]:
    """One sheet as rows, matched on a stripped name.

    Published workbooks put stray whitespace in sheet names often enough that
    an exact lookup is a bug waiting for the next file: Bangladesh's religion
    sheet begins with a space.
    """
    wanted = name.strip().lower()
    for actual in book.sheetnames:
        if actual.strip().lower() == wanted:
            return [list(row) for row in
                    book[actual].iter_rows(values_only=True)]
    raise SystemExit(f"no sheet named {name!r}; the workbook has "
                     f"{', '.join(book.sheetnames)}")


def columns(rows: list[list[Any]]) -> tuple[list[str], list[str]]:
    """The two header rows: field names, and the aliases people read."""
    if len(rows) < 3:
        raise SystemExit("sheet has no data under its two header rows")
    names = ["" if v is None else str(v).strip() for v in rows[0]]
    aliases = ["" if v is None else str(v).strip() for v in rows[1]]
    return names, aliases


def mine(name: str, prefix: str) -> bool:
    """Is this column part of the topic being read?"""
    return not prefix or name.startswith(prefix)


def sexed(names: list[str], prefix: str = "") -> bool:
    """Does this sheet report every group three times, by sex?

    Ethiopia does and the Philippines does not, and the difference is visible
    in the column names rather than declared per country -- a stem that carries
    all three of _B, _F and _M is a sexed sheet.
    """
    stems = {n[:-2] for n in names
             if n not in GEOGRAPHY and n.endswith("_B")
             and mine(n, prefix)}
    if not stems:
        return False
    return all(f"{s}_F" in names and f"{s}_M" in names for s in stems)


def denominator(names: list[str], aliases: list[str],
                prefix: str = "") -> int | None:
    """The column holding everyone the question was asked of, if there is one.

    Found by its alias rather than its name: the Philippines calls it
    ``RLG_HPOP`` / "Household population" and Indonesia's language sheet
    ``LNG_BTOTL`` / "Total population", and inventing a rule that spans both
    from the names alone would be guessing. Ethiopia publishes no such column
    at all, and returning None says so rather than picking a group at random.
    """
    for index, (name, alias) in enumerate(zip(names, aliases)):
        if name in GEOGRAPHY or not name or not mine(name, prefix):
            continue
        if "population" in alias.lower():
            return index
    return None


def groups(names: list[str], aliases: list[str],
           total: int | None, by_sex: bool, prefix: str = "") -> dict[int, str]:
    """Column index -> the label to publish, for every group column."""
    out: dict[int, str] = {}
    for index, (name, alias) in enumerate(zip(names, aliases)):
        if index == total or name in GEOGRAPHY or not name or not alias:
            continue
        if not mine(name, prefix):
            continue
        if by_sex:
            if not name.endswith("_B"):
                continue
            # "Orthodox, both sexes" is the column; "Orthodox" is the group.
            alias = alias.rsplit(",", 1)[0].strip()
        out[index] = alias
    return out


def number(value: Any) -> float | None:
    """A cell as a count of people, or None where the file has none.

    The workbooks write -999 where a figure is unavailable, which is not
    documented in their metadata and is easy to read straight through: it is a
    number, and openpyxl hands it over as one. Four Ethiopian areas carry it
    across all six religions, and taking it at face value would have put a
    negative population on the map.

    Rejected as "not a count" rather than as "-999 specifically", because the
    reason is the stronger one: no census reports a negative number of people,
    so any negative is a marker of some kind whatever its value, and a sentinel
    the Bureau changes to -998 tomorrow would still be caught.
    """
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value) if value >= 0 else None
    if isinstance(value, str):
        text = value.replace(",", "").strip()
        if text.replace(".", "", 1).isdigit():
            return float(text)
    return None


def check_sexes(rows: list[list[Any]], names: list[str], where: str) -> None:
    """Females plus males must be the both-sexes figure, on every row.

    This is the whole reason the sex columns are read at all. They are not a
    dimension this map shows; they are two independent statements about the
    same population, and a file where they disagree is a file this reader has
    misunderstood.
    """
    index = {name: i for i, name in enumerate(names)}
    stems = sorted({n[:-2] for n in names
                    if n not in GEOGRAPHY and n.endswith("_B")})
    off: list[tuple[float, str]] = []
    cells = 0
    for row in rows[2:]:
        where_row = str(row[index["AREA_NAME"]] or "?").strip()
        for stem in stems:
            both = number(row[index[f"{stem}_B"]])
            female = number(row[index[f"{stem}_F"]])
            male = number(row[index[f"{stem}_M"]])
            if both is None or female is None or male is None:
                continue
            cells += 1
            gap_ = (female + male) - both
            if abs(gap_) > 0.5:
                off.append((abs(gap_),
                            f"{where_row} {stem}: {female:,.0f} + {male:,.0f} "
                            f"= {female + male:,.0f} against {both:,.0f} "
                            f"(out by {gap_:+,.0f})"))
    if off:
        off.sort(reverse=True)
        log(f"    {len(off)} of {cells} cells do not reconcile by sex; "
            f"the largest:")
        for _size, line in off[:8]:
            log(f"      {line}")
        raise SystemExit(f"{where}: {len(off)} of {cells} cells where females "
                         f"plus males do not equal the both-sexes figure")
    log(f"    {len(stems)} groups reconcile by sex across {cells} cells")


def area(row: list[Any], names: list[str],
         first_order: int = 1) -> tuple[int | None, str, str]:
    """A row's level, its own name, and its parent's."""
    index = {name: i for i, name in enumerate(names)}
    level = number(row[index["ADM_LEVEL"]])
    name = str(row[index["AREA_NAME"]] or "").strip()
    # The parent is the area at whichever level this country calls admin1,
    # not the level immediately above. Those coincide everywhere the file's
    # second order is this map's -- but Pakistan's do not: the census counts
    # 36 divisions at level 2 and 155 districts at level 3, and it is the
    # districts geoBoundaries draws, directly under the provinces. Taking the
    # level above would scope a district inside a division no boundary file
    # has, which is a parent that can never match rather than one that
    # sometimes does.
    parent = ""
    above = f"ADM{first_order}_NAME"
    if level is not None and level > first_order and above in index:
        parent = str(row[index[above]] or "").strip()
    return (int(level) if level is not None else None, name, parent)


def read(book, country: Country,
         topic: Topic) -> dict[tuple[str, str], dict[str, Any]]:
    """One topic's sheet, keyed by parent and name rather than name alone.

    Ethiopia has a North Shewa in Amhara and another in Oromia -- geoBoundaries
    distinguishes them as "North Shewa(R3)" and "North Shewa(R4)" -- and keying
    on the name alone let one silently overwrite the other. The count was the
    only sign: 92 zones written where the sheet lists 93.
    """
    rows = sheet_rows(book, topic.sheet)
    names, aliases = columns(rows)
    by_sex = sexed(names, topic.prefix)
    if topic.denominator:
        if topic.denominator not in names:
            raise SystemExit(
                f"{country.iso3} {topic.sheet}: no column named "
                f"{topic.denominator!r} to use as the denominator")
        total = names.index(topic.denominator)
    else:
        total = denominator(names, aliases, topic.prefix)
    found = groups(names, aliases, total, by_sex, topic.prefix)
    if not found:
        raise SystemExit(
            f"{country.iso3} {topic.sheet}: no group columns to read"
            + (f" under the prefix {topic.prefix!r}" if topic.prefix
               else "; every column is geography"))
    log(f"  {topic.sheet}"
        + (f" [{topic.prefix}]" if topic.prefix else "")
        + f": {len(rows) - 2} rows, {len(found)} groups"
        + (", by sex" if by_sex else "")
        + (f", denominator {aliases[total]!r}" if total is not None
           else ", no published denominator"))
    if by_sex:
        check_sexes(rows, names, f"{country.iso3} {topic.sheet}")

    out: dict[tuple[str, str], dict[str, Any]] = {}
    skipped = 0
    empty: list[str] = []
    whole: dict[str, float] = {}
    # Children of an area whose own row is blank, accumulated as they go past.
    # Every child contributes its published total whether or not it has any
    # figures, because a rayon the source lists and leaves empty is part of the
    # oblast's population and leaving it out of the denominator would hide the
    # hole instead of drawing it.
    kids: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"counts": defaultdict(float), "published": 0.0,
                 "listed": 0, "figured": 0, "blank": []})
    # Which ADM level this country calls its first order. Every country but
    # Pakistan says 1, and Pakistan's districts hang off the provinces even
    # though the census puts a level of divisions between them.
    first_order = next((n for n, where in sorted(country.levels.items())
                        if where == "admin1"), 1)
    for row in rows[2:]:
        level, name, parent = area(row, names, first_order)
        if level == 0 and name:
            # The country's own row: the control every other row is checked
            # against. Not a record on this map -- the country already has one
            # -- but the only independent statement of what the parts add to.
            whole = {label: value for index, label in found.items()
                     if (value := number(row[index])) is not None and value > 0}
            continue
        if not name or name.upper() == "NO NAME":
            # "NO NAME" is the workbook's own placeholder for an area it could
            # not label. It is not a place, and giving it a record would put an
            # unnamed shape on the map with figures attached.
            skipped += 1
            continue
        counts = {label: value for index, label in found.items()
                  if (value := number(row[index])) is not None and value > 0}
        # Children are added up before the level filter, so a level can feed a
        # parent without appearing on the map itself. Ukraine needs exactly
        # that: its rayons are the only place the figures exist, and its
        # oblasts are the only level that joins.
        if country.sum_into is not None and level == country.sum_into + 1 \
                and parent:
            kid = kids[parent]
            kid["listed"] += 1
            if (its_own := number(row[total]) if total is not None else None):
                kid["published"] += its_own
            if counts:
                kid["figured"] += 1
                for label, value in counts.items():
                    kid["counts"][label] += value
            else:
                kid["blank"].append(name)
        if level not in country.levels:
            skipped += 1
            continue
        if not counts:
            # An area the file lists and has no figures for. Named rather than
            # dropped in silence: three Ethiopian first-order divisions arrive
            # this way, and their absence is what made the regions add up to
            # three million fewer people than the census counted.
            empty.append(name)
            continue
        published = number(row[total]) if total is not None else None
        key = (parent, name)
        if key in out:
            raise SystemExit(
                f"{country.iso3} {topic.sheet}: two rows for {name!r} under "
                f"{parent!r}; the sheet's own key is not unique")
        out[key] = {"level": level, "parent": parent, "name": name,
                    "counts": counts, "published": published,
                    "summed": sum(counts.values())}
    # Build the blank parents from the children just gathered. After the loop,
    # so that an area publishing its own figures is already in `out` and is
    # left alone.
    built = []
    for name, kid in sorted(kids.items()):
        key = ("", name)
        if key in out or not kid["counts"]:
            continue
        counts = dict(kid["counts"])
        out[key] = {"level": country.sum_into, "parent": "", "name": name,
                    "counts": counts, "published": kid["published"] or None,
                    "summed": sum(counts.values()),
                    "built_from": kid["figured"], "listed": kid["listed"]}
        built.append((name, kid))
    for name, kid in built:
        missing = kid["listed"] - kid["figured"]
        if missing:
            log(f"    {name}: summed from {kid['figured']} of "
                f"{kid['listed']} children; {missing} carry no figures "
                f"({', '.join(sorted(kid['blank'])[:3])}"
                + (" ..." if missing > 3 else "") + ")")
    if built:
        log(f"    {len(built)} areas at level {country.sum_into} built by "
            f"adding up the level below")
    if skipped:
        log(f"    {skipped} rows skipped: another level, or no area name")
    # An area that was blank and has since been built is not missing, so it is
    # not named here -- otherwise Ukraine would report 25 oblasts with nothing
    # in the same run that fills all 25.
    still_empty = sorted(set(empty) - {name for name, _kid in built})
    if still_empty:
        log(f"    {len(still_empty)} areas listed with no figures at all: "
            + ", ".join(still_empty[:6])
            + (" ..." if len(still_empty) > 6 else ""))
    if whole:
        for level, layer in sorted(country.levels.items()):
            parts = sum(sum(r["counts"].values())
                        for r in out.values() if r["level"] == level)
            if not parts:
                continue
            total = sum(whole.values())
            short = total - parts
            log(f"    {layer}: {parts:,.0f} against the country's own "
                f"{total:,.0f}"
                + (f" -- {short:,.0f} fewer ({100 * short / total:.1f}%)"
                   if abs(short) > 0.5 else " -- exactly"))
    return out


def check_total(country: Country, topic: Topic,
                areas: dict[str, dict[str, Any]]) -> None:
    """Where the sheet publishes a denominator, how near do the groups come?

    Falling short is not by itself an error. The published total is everyone
    the question was asked of, and a census may decline to publish a group too
    small to disclose -- so the remainder is people whose answer this table does
    not show, which the map draws as an explicit unaccounted share rather than
    normalising away. What would be an error is falling short *everywhere*, or
    by a lot, and those are refused.
    """
    checked = 0
    short: list[tuple[float, str]] = []
    worst = 0.0
    for (_parent, name), row in areas.items():
        if row["published"] is None or row["published"] <= 0:
            continue
        checked += 1
        gap_ = (row["published"] - row["summed"]) / row["published"]
        if abs(gap_) > NOTABLE:
            worst = max(worst, abs(gap_))
            short.append((abs(gap_),
                          f"{name}: groups come to {row['summed']:,.0f} "
                          f"against a published {row['published']:,.0f} "
                          f"({gap_:+.1%})"))
    if not checked:
        return
    if not short:
        log(f"    every one of {checked} areas reaches its published total")
        return
    short.sort(reverse=True)
    log(f"    {len(short)} of {checked} areas do not reach their published "
        f"total; the map shows the remainder as unaccounted:")
    for _size, line in short[:6]:
        log(f"      {line}")
    if worst > country.refuse_area:
        raise SystemExit(f"{country.iso3} {topic.sheet}: an area is "
                         f"{worst:.1%} short of its own published total, too "
                         f"much to be a suppressed group")
    if len(short) / checked > country.refuse_share:
        raise SystemExit(f"{country.iso3} {topic.sheet}: {len(short)} of "
                         f"{checked} areas fall short, which is a reader that "
                         f"has misunderstood the sheet rather than a source "
                         f"with holes in it")


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    # Deliberately unconstrained: --discover takes ISO3 codes for countries
    # that are not configured yet, which is the whole point of it. The read
    # path checks the code against COUNTRIES itself, and says what it knows.
    ap.add_argument("--country", action="append", default=None,
                    metavar="ISO3",
                    help=f"one of {', '.join(sorted(COUNTRIES))} or all; "
                         f"with --discover, any ISO3 code")
    ap.add_argument("--out")
    ap.add_argument("--discover", action="store_true",
                    help="report which countries in the series carry a USCB "
                         "workbook, instead of reading any")
    ap.add_argument("--inspect", action="append", metavar="DATASET",
                    help="print the levels, groups, areas and census year of "
                         "one HDX dataset's workbook, which is what writing a "
                         "config needs and the catalogue does not carry")
    ap.add_argument("--sheet", action="append",
                    help="with --inspect, only sheets whose name contains this")
    ap.add_argument("--limit", type=int, default=25,
                    help="with --discover, how many datasets to report "
                         "(0 for all)")
    ap.add_argument("--sheets", action="store_true",
                    help="with --discover, open each workbook and list its "
                         "sheets (slow; the topics are only in the file)")
    args = ap.parse_args()

    if args.inspect:
        for dataset in args.inspect:
            log(f"\n{'=' * 70}")
            inspect(dataset, args.sheet or [])
        return 0
    if args.discover:
        return discover(args.limit, args.sheets)

    import openpyxl

    wanted = [c for c in (args.country or ["all"])]
    chosen = (sorted(COUNTRIES) if "all" in wanted else wanted)

    unknown = [c for c in chosen if c not in COUNTRIES]
    if unknown:
        raise SystemExit(
            f"no configuration for {', '.join(unknown)}. Configured: "
            f"{', '.join(sorted(COUNTRIES))}.\n"
            f"Try --discover --country {unknown[0]} --sheets to see whether "
            f"the series carries a workbook for it.")

    for iso3 in chosen:
        country = COUNTRIES[iso3]
        source = workbook_url(country.dataset)
        log(f"{country.name}: {source}")
        blob = http_get(source, binary=True, cache_dir=RAW / "uscb")
        log(f"  {len(blob):,} bytes")
        book = openpyxl.load_workbook(io.BytesIO(blob), read_only=True,
                                      data_only=True)

        fields: dict[str, dict[str, dict[str, Any]]] = {}
        for topic in country.topics:
            areas = read(book, country, topic)
            check_total(country, topic, areas)
            fields[topic.field] = areas
        book.close()

        every = sorted({key for areas in fields.values() for key in areas})
        records: list[dict[str, Any]] = []
        for key in every:
            parent, name = key
            any_row = next(a[key] for a in fields.values() if key in a)
            level = country.levels[any_row["level"]]
            values: dict[str, Any] = {}
            # One source entry per topic that produced a figure, rather than
            # one for the country naming every field. Burma's ethnicity comes
            # from a different publication and a different year than the rest
            # of its workbook, so a single citation covering both would be
            # wrong about one of them whichever way it was written.
            cites: list[dict[str, Any]] = []
            for topic in country.topics:
                row = fields[topic.field].get(key)
                if not row:
                    continue
                published = shares(
                    row["counts"], total=row["published"] or row["summed"])
                if not country.counts_are_people:
                    published = [{"group": g["group"], "pct": g["pct"]}
                                 for g in published]
                values[topic.field] = published or gap(NOT_AVAILABLE)
                values[f"{topic.field}_year"] = topic.year or country.year
                values[f"{topic.field}_note"] = topic.note or country.note
                cites.append({"field": topic.field,
                              "name": topic.source or country.source,
                              "url": dataset_url(country.dataset),
                              "license": country.licence})
            def slug(text: str) -> str:
                return "".join(c if c.isalnum() else "-"
                               for c in text.lower()).strip("-")
            # The parent is part of the id for the same reason it is part of
            # the key: two zones may share a name and are not the same place.
            ident = f"{iso3}-{slug(parent)}-{slug(name)}" if parent \
                else f"{iso3}-{slug(name)}"
            records.append(record(
                ident, name.title(), level=level, parent=iso3,
                parent_name=parent.title() or None,
                aliases=list(country.aliases.get(name.title(), ())),
                # The parent's aliases travel with the row for the same reason
                # its name does. Without them a zone can only be scoped against
                # a region the boundary file spells differently, which for
                # Ethiopia is every region with a macron in it -- so no zone
                # was ever matched *inside* its region, and the two North
                # Shewas, one in Amhara and one in Oromia, were both refused as
                # ambiguous when the region each belongs to tells them apart.
                parent_aliases=list(country.aliases.get(parent.title(), ())),
                # Declared absence. A row that says it has no boundary is a
                # visible gap; the same row left to the matcher is whatever
                # shape it happens to reach.
                # A first-order area's parent is the country, and naming it
                # keeps every declaration a (parent, name) pair: a name alone
                # is not an address one level down, where the Philippines has
                # a Quezon province and a Quezon city 130 km apart.
                no_shape=(parent.title() or country.name,
                          name.title()) in country.no_shape or None,
                sources=cites,
                **values))

        out = args.out or PROCESSED / country.out
        write_json(out, records)
        counts = {lvl: sum(1 for r in records if r["level"] == lvl)
                  for lvl in sorted(country.levels.values())}
        log(f"  {len(records)} areas: "
            + ", ".join(f"{n} {lvl}" for lvl, n in counts.items()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
