#!/usr/bin/env python3
"""United States -- Census Bureau ACS 5-year API (states and counties).

Race/ethnicity comes from **B03002** (Hispanic or Latino origin by race), not
B02001, because only B03002 makes the Hispanic-origin question orthogonal to
race the way the published "White, non-Hispanic" figures do.  Language uses
**C16001** (the collapsed version of B16001, which is far smaller over 3,143
counties).  Median age and sex ratio come from the **DP05** profile.

Religion is *not* in the census: the US census has been barred from asking a
mandatory religion question since 1976 (13 U.S.C. 221(c)).  The county-level
substitute is the 2020 U.S. Religion Census (ASARB, distributed by ARDA), which
counts *adherents reported by 372 religious bodies* -- 161,224,088 people, about
48.6% of the 2020 population -- and is therefore not comparable with the
self-identification percentages used everywhere else in this dataset.

The 372 individual bodies it reports are collapsed into traditions -- Catholic,
Protestant, Orthodox Christian, Latter-day Saints, Judaism, Islam, Buddhism,
Hinduism and so on -- because a denominational breakdown is neither mappable nor
comparable with the broad census categories used elsewhere here.  Output is
always labelled ``religion_basis: adherents`` and carries the publisher's
suggested citation.

An API key is optional below 500 calls/day; set ``CENSUS_API_KEY`` to lift that.

Usage:
    python -m scripts.fetch_census.us_acs --level county --year 2022
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from typing import Any

from ._shared import (
    NOT_AVAILABLE, NOT_COLLECTED, PROCESSED, RAW, gap, http_get, http_json, log,
    measure, record, shares, write_json,
)

BASE = "https://api.census.gov/data/{year}/acs/acs5"
# DP* variables are served by the data-profile endpoint, not the detailed
# tables -- asking acs/acs5 for DP05_0018E gets a plain-text error, not JSON.
BASE_PROFILE = "https://api.census.gov/data/{year}/acs/acs5/profile"

# B03002 lines that partition the population exactly once.
RACE_LINES = {
    "B03002_003E": "White (non-Hispanic)",
    "B03002_004E": "Black or African American (non-Hispanic)",
    "B03002_005E": "American Indian and Alaska Native (non-Hispanic)",
    "B03002_006E": "Asian (non-Hispanic)",
    "B03002_007E": "Native Hawaiian and Other Pacific Islander (non-Hispanic)",
    "B03002_008E": "Some other race (non-Hispanic)",
    "B03002_009E": "Two or more races (non-Hispanic)",
    "B03002_012E": "Hispanic or Latino (any race)",
}
RACE_TOTAL = "B03002_001E"

LANGUAGE_LINES = {
    "C16001_002E": "English only",
    "C16001_003E": "Spanish",
    "C16001_006E": "French, Haitian, or Cajun",
    "C16001_009E": "German or other West Germanic",
    "C16001_012E": "Russian, Polish, or other Slavic",
    "C16001_015E": "Other Indo-European",
    "C16001_018E": "Korean",
    "C16001_021E": "Chinese (incl. Mandarin, Cantonese)",
    "C16001_024E": "Vietnamese",
    "C16001_027E": "Tagalog (incl. Filipino)",
    "C16001_030E": "Other Asian and Pacific Island",
    "C16001_033E": "Arabic",
    "C16001_036E": "Other and unspecified",
}
LANGUAGE_TOTAL = "C16001_001E"

PROFILE_LINES = {"DP05_0018E": "median_age", "DP05_0004E": "sex_ratio_m_per_100f"}


def query(year: int, get: list[str], geo: str, key: str | None,
          base: str = BASE) -> list[dict[str, str]]:
    params = [f"get=NAME,{','.join(get)}", f"for={geo}"]
    if key:
        params.append(f"key={key}")
    url = f"{base.format(year=year)}?" + "&".join(params)
    text = http_get(url, timeout=180)
    assert isinstance(text, str)
    # The Census API used to allow 500 anonymous calls a day; it now returns a
    # "Missing Key" HTML page with HTTP 200 when no key is sent. Turn that into
    # an instruction instead of a JSON traceback.
    if "Missing Key" in text[:400]:
        raise SystemExit(
            "us_acs: api.census.gov now requires an API key for every request. "
            "Request a free key at https://api.census.gov/data/key_signup.html "
            "and set it as the CENSUS_API_KEY environment variable (in CI: a "
            "repository secret of the same name).")
    import json as _json
    rows = _json.loads(text.lstrip("\ufeff"))
    header, *body = rows
    return [dict(zip(header, row)) for row in body]


def as_float(value: str | None) -> float | None:
    """ACS uses negative sentinels (-666666666) for suppressed cells."""
    if value in (None, "", "null"):
        return None
    try:
        num = float(value)
    except ValueError:
        return None
    return None if num <= -666666 else num


def geoid(row: dict[str, str], level: str) -> str:
    return row["state"] + (row.get("county", "") if level == "county" else "")


def fetch(level: str, year: int, key: str | None) -> list[dict[str, Any]]:
    geo = "county:*" if level == "county" else "state:*"
    src = f"U.S. Census Bureau, ACS {year} 5-year estimates"

    race = query(year, [RACE_TOTAL, *RACE_LINES], geo, key)
    lang = {geoid(r, level): r for r in query(year, [LANGUAGE_TOTAL, *LANGUAGE_LINES], geo, key)}
    prof = {geoid(r, level): r
            for r in query(year, list(PROFILE_LINES), geo, key, base=BASE_PROFILE)}

    out: list[dict[str, Any]] = []
    for row in race:
        gid = geoid(row, level)
        total = as_float(row.get(RACE_TOTAL))
        counts = {label: as_float(row.get(code)) for code, label in RACE_LINES.items()}
        counts = {k: v for k, v in counts.items() if v is not None}

        lrow = lang.get(gid, {})
        lcounts = {label: as_float(lrow.get(code)) for code, label in LANGUAGE_LINES.items()}
        lcounts = {k: v for k, v in lcounts.items() if v}

        prow = prof.get(gid, {})
        median = as_float(prow.get("DP05_0018E"))
        ratio = as_float(prow.get("DP05_0004E"))

        # ACS names a county "Autauga County, Alabama". The state half is what
        # separates the thirty-one counties called Washington from each other, so
        # it is passed through for matching rather than thrown away.
        name = row["NAME"]
        state_name = name.split(",")[-1].strip() if level == "county" and "," in name else None

        out.append(record(
            f"USA-{gid}",
            name,
            level="admin1" if level == "state" else "admin2",
            parent="USA" if level == "state" else f"USA-{row['state']}",
            parent_name=state_name,
            codes={"geoid": gid, "fips_state": row["state"],
                   "fips_county": row.get("county")},
            population=measure(int(total), year=year, source=src) if total else gap(NOT_AVAILABLE),
            median_age=measure(median, unit="years", year=year, source=src) if median else gap(NOT_AVAILABLE),
            sex_ratio=(measure(round(ratio * 10), unit="males_per_1000_females",
                               year=year, source=src) if ratio else gap(NOT_AVAILABLE)),
            ethnicity=shares(counts, total=total) or gap(NOT_AVAILABLE),
            ethnicity_note=("US Census race and Hispanic-origin categories (ACS table B03002). "
                            "Not comparable with other countries' ethnicity classifications."),
            language=shares(lcounts, total=as_float(lrow.get(LANGUAGE_TOTAL))) or gap(NOT_AVAILABLE),
            language_note="Language spoken at home, population 5 years and over (ACS table C16001).",
            religion=gap(NOT_COLLECTED, CENSUS_BARRED + " Any figures shown come "
                         "instead from the 2020 U.S. Religion Census (ASARB), which "
                         "counts adherents reported by religious bodies rather than "
                         "asking people."),
            sources=[{"field": "ethnicity/language/population", "name": src,
                      "url": BASE.format(year=year), "license": "Public domain (U.S. Government work)"}],
        ))
    return out


# The citation the publisher asks for, verbatim, recorded on every record the
# study touches. Note the year: the workbook's own Copyright sheet says 2022,
# the suggested citation on usreligioncensus.org says 2023. The publisher's
# wording is what goes on the record.
# Checked in under data/raw/us/: there is no API for this study, and the US
# census is barred from asking the question, so without the workbook the build
# cannot re-derive religion for a single county.
RELIGION_FILE = RAW / "us" / "2020_USRC_Group_Detail.xlsx"

RELIGION_CITATION = (
    "Clifford Grammich, Erica Dollhopf, Mary Gautier, Richard Houseal, "
    "Dale E. Jones, Alexei Krindatch, Richie Stanley, and Scott Thumma. 2023. "
    "2020 U.S. Religion Census: Religious Congregations & Membership Study. "
    "Association of Statisticians of American Religious Bodies."
)

# Every sheet in the workbook carries a row holding the whole country, and each
# one hides it differently: the county sheet keys it FIPS = "Total", the state
# sheet StateCode = "Totals", the nation sheet Group Code = "Totals". None of
# them names a group. Summing a column without excluding it doubles the United
# States exactly -- 322,019,032 adherents against a true 161,009,516, 97% of the
# population religiously adherent.
#
# Nothing internal to the table catches that. Every county's own shares stay
# correct and still reproduce the percentages the file prints beside them,
# exactly as they did for the ABS "Christianity Total" rows and the India C-16
# group codes. It is visible only against a total the detail rows did not
# produce, which is why these rows are read and used as the control rather than
# skipped and forgotten. The singular/plural difference is the whole reason this
# is a set: an exact match on "total" silently misses the state sheet.
NATIONAL_BLOCK = {"total", "totals"}

LEVELS = {
    "county": {"sheet": "2020 Group by County", "key": ("fips",), "width": 5,
               "share": ("adherents as % of total population",)},
    "state": {"sheet": "2020 Group by State", "key": ("statecode",), "width": 2,
              "share": ("adherents as % of population",)},
}

# The study reports 372 individual religious bodies, which is far too fine to
# put on a map and not comparable with the broad census categories every other
# country here uses. They are collapsed into traditions.
#
# The mapping is by exact name and only for what is *not* Protestant; anything
# unlisted falls through to Protestant, and the adapter prints the largest
# groups it defaulted so that a body added in a later release is visible rather
# than silently absorbed. Matching on keywords instead would be wrong in both
# directions: the Orthodox Presbyterian Church and the Orthodox Mennonite Church
# are Protestant, and the Polish National Catholic Church is not Roman Catholic.
DEFAULT_TRADITION = "Protestant"
TRADITIONS: dict[str, str] = {
    "Catholic Church": "Catholic",
    "Church of Jesus Christ of Latter-day Saints": "Latter-day Saints",
    "Community of Christ": "Latter-day Saints",
    "Jehovah's Witnesses": "Jehovah's Witnesses",
    "Muslim Estimate": "Islam",
    # Judaism is reported by movement; the map shows the religion.
    "Orthodox Judaism": "Judaism",
    "Reform Judaism": "Judaism",
    "Conservative Judaism": "Judaism",
    "Reconstructionist Judaism": "Judaism",
    "Independent Judaism": "Judaism",
    "Chabad Judaism": "Judaism",
    "Hindu Temples": "Hinduism",
    "Hindu Yoga and Meditation": "Hinduism",
    "Vedanta Society": "Hinduism",
    "Mahayana Buddhist": "Buddhism",
    "Theravada Buddhist": "Buddhism",
    "Vajarayana Buddhist": "Buddhism",
    # Eastern and Oriental Orthodox, by jurisdiction.
    "Greek Orthodox Archdiocese of America": "Orthodox Christian",
    "Coptic Orthodox Church": "Orthodox Christian",
    "Ethiopian Orthodox": "Orthodox Christian",
    "Eritrean Orthodox": "Orthodox Christian",
    "Orthodox Church in America": "Orthodox Christian",
    "Antiochian Orthodox Christian Archdiocese of North America, The": "Orthodox Christian",
    "Serbian Orthodox Church in North America": "Orthodox Christian",
    "Armenian Church of North America (Catholicosate of Etchmiadzin)": "Orthodox Christian",
    "Armenian Apostolic Church of America (Catholicosate of Cilicia)": "Orthodox Christian",
    "Russian Orthodox Church Outside of Russia": "Orthodox Christian",
    "Patriarchal Parishes of the Russian Orthodox Church in the USA": "Orthodox Christian",
    "Malankara Orthodox Syrian Church": "Orthodox Christian",
    "Malankara Archdiocese of the Syrian Orthodox Church in North America": "Orthodox Christian",
    "Syriac Orthodox Church of Antioch": "Orthodox Christian",
    "Macedonian Orthodox Church: American Diocese": "Orthodox Christian",
    "Ukrainian Orthodox Church of the USA": "Orthodox Christian",
    "Romanian Orthodox Archdiocese in Americas": "Orthodox Christian",
    "American Carpatho-Russian Orthodox Diocese": "Orthodox Christian",
    "Bulgarian Eastern Orthodox Diocese of the USA, Canada and Australia": "Orthodox Christian",
    "Albanian Orthodox Diocese of America": "Orthodox Christian",
    "Georgian Orthodox Parishes in the United States": "Orthodox Christian",
    "Belarusan Autocephalous Orthodox Church": "Orthodox Christian",
    "Church of the Genuine Orthodox Christians": "Orthodox Christian",
    "Holy Orthodox Church in North America": "Orthodox Christian",
    "Syro-Russian Orthodox Catholic Church": "Orthodox Christian",
    # Old Catholic and independent Catholic bodies, in communion with none of
    # the above and not counted as Roman Catholic.
    "Polish National Catholic Church": "Other Christian",
    "North American Old Roman Catholic Church": "Other Christian",
    "Orthodox Old Roman Catholic Communion": "Other Christian",
    "Ecumenical Catholic Communion": "Other Christian",
    "Ecumenical Catholic Church": "Other Christian",
    "United Catholic Church": "Other Christian",
    "Liberal Catholic Church": "Other Christian",
    "Catholic Apostolic Church in North America": "Other Christian",
    "Swedenborgian Church": "Other Christian",
    "Union of Messianic Jewish Congregations": "Other Christian",
    "Association of Messianic Congregations": "Other Christian",
    # Reported as congregations with no adherent estimate, so these contribute
    # nothing to any share; they are listed so the classification is complete
    # and so they land correctly if a later release does estimate them.
    "Baha'i Faith USA": "Other religions",
    "American Sikh Council": "Other religions",
    "Jain": "Other religions",
    "Zoroastrian": "Other religions",
    "Shinto": "Other religions",
    "Tao": "Other religions",
    "Unitarian Universalist Association of Congregations": "Other religions",
    "National Spiritualist Association of Churches": "Other religions",
}


# The reason the United States has no government religion figure at any level of
# geography, stated once. It prefixes every gap this adapter emits, because it
# is the fact that makes the rest of the sentence necessary.
CENSUS_BARRED = ("The U.S. census may not ask a mandatory religion question "
                 "(13 U.S.C. 221(c)), so no government figures exist at any "
                 "level.")

# Why a US area can have no figure even with the workbook in hand. Naming the
# reason is the whole point: "no data" and "the study does not cover this place"
# and "nobody reported a congregation here" are three different facts, and only
# the last is about the place itself.
UNMATCHED_REASONS = {
    "72": "The 2020 study covers the 50 states and the District of Columbia. "
          "Puerto Rico is outside its frame, so no municipio has an adherent "
          "count from it either.",
    "09": "Connecticut replaced its eight counties with nine planning regions in "
          "2022. The 2020 study reports the old counties, so its figures cannot "
          "be placed on this geography without inventing a way to split them.",
}
UNMATCHED_DEFAULT = (
    "No religious body reporting to the 2020 study had a congregation here. That "
    "is an absence of reported adherents, not a count of zero believers.")


def read_group_detail(path: Path, level: str) -> tuple[dict[str, dict[str, float]],
                                                       dict[str, float], float]:
    """Adherents by area and religious body, area populations, and the control.

    Handles the ASARB workbook directly (.xlsx) and a CSV export of one of its
    sheets. Returns per-area counts keyed by the body's own name, the 2020
    population the file itself implies for each area, and the national total
    carried by the whole-country row.
    """
    spec = LEVELS[level]
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = book[spec["sheet"]] if spec["sheet"] in book.sheetnames else book.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = [str(c or "") for c in next(rows)]
        records: Any = (dict(zip(header, r)) for r in rows)
    else:
        import csv
        records = csv.DictReader(path.open(newline="", encoding="utf-8-sig"))

    def field(row: dict, *names: str) -> Any:
        for name in names:
            for key in row:
                if key and key.strip().lower() == name:
                    return row[key]
        return None

    areas: dict[str, dict[str, float]] = {}
    populations: dict[str, float] = {}
    national = 0.0
    for row in records:
        code = str(field(row, *spec["key"]) or "").strip()
        if not code:
            continue
        try:
            adherents = float(field(row, "adherents", "adherent"))
        except (TypeError, ValueError):
            continue          # blank means not reported, which is not zero

        # Before anything else: the whole-country row names no group, so a
        # reader that requires one skips it silently and loses the only figure
        # the detail rows can be checked against.
        if code.lower() in NATIONAL_BLOCK:
            national += adherents
            continue

        group = str(field(row, "group name", "grpname") or "").strip()
        if not group:
            continue
        code = code.zfill(spec["width"])
        areas.setdefault(code, {})[group] = adherents
        # The file prints each count as a share of its area's population, so the
        # denominator it used can be recovered rather than assumed. Taking it
        # from the file keeps the shares equal to the published ones; the ACS
        # population on the record is a different year and would not.
        try:
            share = float(field(row, *spec["share"]))
        except (TypeError, ValueError):
            share = 0.0
        if share > 0 and adherents > 0:
            implied = adherents / share
            known = populations.setdefault(code, implied)
            # Every row of an area must imply the same denominator. If the file
            # ever computed its percentages against something that varies by
            # row, taking the population from the first row would leave every
            # other share on that area wrong by a factor nothing would show.
            if abs(implied - known) > max(1.0, known * 1e-6):
                raise SystemExit(
                    f"{level} {code}: rows imply different populations "
                    f"({known:,.0f} and {implied:,.0f}); the share column is not "
                    "on a single denominator and cannot be inverted")
    return areas, populations, national


def check_national(areas: dict[str, dict[str, float]], national: float, level: str) -> None:
    """The areas must add up to the whole-country row, or the read is wrong.

    This is the check that catches reading that row as an area. It compares
    against a figure the detail rows did not produce, so unlike a
    shares-add-to-100% test it cannot be satisfied by double counting.
    """
    if not national:
        raise SystemExit(
            f"no whole-country row (key in {sorted(NATIONAL_BLOCK)}) in the {level} "
            "sheet. Either the layout changed or it was filtered out -- without it "
            "the detail rows have nothing independent to reconcile against, and a "
            "doubled country reads as a valid table.")
    total = sum(sum(g.values()) for g in areas.values())
    drift = abs(total - national) / national
    print(f"  {level}: {total:,.0f} vs whole-country row {national:,.0f} "
          f"({drift:.4%} apart, {len(areas)} areas)")
    if drift > 0.005:
        raise SystemExit(f"{level} adherents are {drift:.2%} from the whole-country row")


def to_traditions(groups: dict[str, float]) -> dict[str, float]:
    """Collapse individual religious bodies into the traditions the map shows."""
    out: dict[str, float] = {}
    for name, adherents in groups.items():
        tradition = TRADITIONS.get(name, DEFAULT_TRADITION)
        out[tradition] = out.get(tradition, 0.0) + adherents
    return out


def report_defaults(areas: dict[str, dict[str, float]]) -> None:
    """Name the largest bodies that fell through to Protestant.

    Everything unlisted defaults, so a body added by a later release -- or a
    renamed one -- would be absorbed without complaint. Printing the top of that
    list on every run is what makes the classification auditable.
    """
    totals: dict[str, float] = {}
    for groups in areas.values():
        for name, adherents in groups.items():
            if name not in TRADITIONS:
                totals[name] = totals.get(name, 0.0) + adherents
    top = sorted(totals.items(), key=lambda kv: kv[1], reverse=True)[:5]
    print(f"  {len(totals)} bodies defaulted to {DEFAULT_TRADITION}; largest: "
          + ", ".join(f"{n} {v:,.0f}" for n, v in top))


def attach_religion(records: list[dict[str, Any]], path: Path, level: str) -> None:
    """Merge a 2020 U.S. Religion Census extract (ASARB, ARDA dataset RCMSCY20).

    Opt-in: the study is copyright ASARB and the map carries it by the
    publisher's suggested citation, which is recorded on every record it
    touches.
    """
    areas, populations, national = read_group_detail(path, level)
    check_national(areas, national, level)
    report_defaults(areas)

    matched = 0
    for rec in records:
        gid = rec.get("codes", {}).get("geoid", "")
        groups = areas.get(gid)
        pop = populations.get(gid)
        if not groups or not pop:
            rec["religion"] = gap(NOT_COLLECTED, CENSUS_BARRED + " " +
                                  UNMATCHED_REASONS.get(gid[:2], UNMATCHED_DEFAULT))
            continue
        matched += 1
        rec["religion"] = shares(to_traditions(groups), total=pop)
        rec["religion_basis"] = "adherents"
        rec["religion_note"] = (
            "2020 U.S. Religion Census (ASARB): adherents reported by 372 religious "
            "bodies, grouped into traditions and expressed as a share of the 2020 "
            "census population. This is not self-identification and does not sum to "
            "100% -- the study reached about 48.6% of the population nationally, and "
            "the remainder is uncounted rather than unaffiliated. Not comparable "
            "with the census religion figures used for other countries.")
        rec["sources"].append({"field": "religion", "name": RELIGION_CITATION,
                               "url": "https://www.usreligioncensus.org/",
                               "license": "Copyright ASARB; used with the "
                                          "publisher's suggested citation"})
    print(f"  religion attached to {matched} of {len(records)} records")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--level", default="county", choices=["state", "county"])
    ap.add_argument("--year", type=int, default=2022)
    ap.add_argument("--religion-file", type=Path, default=RELIGION_FILE,
                    help="2020 U.S. Religion Census workbook or sheet export "
                         "(ASARB / ARDA RCMSCY20); read at the chosen --level")
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()

    key = os.environ.get("CENSUS_API_KEY")
    log(f"us_acs: ACS {args.year} 5-year, level={args.level}"
        + ("" if key else " (no CENSUS_API_KEY set; the API now rejects keyless requests)"))
    records = fetch(args.level, args.year, key)
    if args.religion_file and args.religion_file.exists():
        attach_religion(records, args.religion_file, args.level)
    else:
        # Not an error: without the workbook the records keep their
        # not_collected marker, which is the truthful state of a US religion
        # figure. Saying so beats a silent skip.
        print(f"  no religion workbook at {args.religion_file}; "
              "religion stays not_collected")
    out = args.out or PROCESSED / f"us_{args.level}.json"
    write_json(out, records)
    log(f"  {len(records)} {args.level} records")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
