#!/usr/bin/env python3
"""India -- Census 2011 religion and population, by district and state.

India is the one major country in this pipeline with no statistics API. The
Registrar General publishes table C-01 (population by religious community) as
per-state XLSX workbooks through the censusindia.gov.in NADA catalogue, which
cannot be automated.

So the default path reads a **district-level CSV extract of the 2011 primary
census abstract** and aggregates it upward. Because that extract is a community
redistribution rather than an official endpoint, it is not taken on trust: every
run re-checks it against the figures the Registrar General published, and
refuses to emit anything if they drift (see ``NATIONAL_CONTROLS`` and
``validate``). The current file reproduces the official national totals exactly
-- 1,210,854,977 people, Hindu 79.80%, Muslim 14.23%, Christian 2.30%, Sikh
1.72%, Buddhist 0.70%, Jain 0.37% -- and its per-state aggregates reproduce the
independently hand-compiled rows in ``data/curated/admin1_seed.json`` to within
rounding.

Two standing caveats the app displays with every Indian figure:

* The reference year is **2011**. The next census was postponed repeatedly and
  had not been conducted at the time of writing, so India's subnational
  demographics are more than a decade old.
* India does not collect ethnicity. Scheduled Caste and Scheduled Tribe shares
  (a constitutional-schedule classification, not an ethnic one) are collected
  instead, and are emitted here as their own field rather than folded into
  "ethnicity".

Mother tongue (table C-16) is *not* in this extract, so language stays an
explicit gap.

Usage:
    python -m scripts.fetch_census.india_census --level district
    python -m scripts.fetch_census.india_census --level state
    python -m scripts.fetch_census.india_census --input data/raw/india   # XLSX
"""

from __future__ import annotations

import argparse
import collections
import csv
import io
from pathlib import Path
from typing import Any

from ._shared import (
    NOT_AVAILABLE, NOT_COLLECTED, PROCESSED, RAW, gap, http_get, log, measure,
    record, shares, write_json,
)

CATALOG = "https://censusindia.gov.in/nada/index.php/catalog"
CSV_URL = ("https://raw.githubusercontent.com/nishusharma1608/"
           "India-Census-2011-Analysis/master/india-districts-census-2011.csv")

SOURCE = "Census of India 2011, table C-01 (Registrar General & Census Commissioner)"
SOURCE_NOTE = ("District-level extract of the 2011 primary census abstract. Validated "
               "on every run against the Registrar General's published national totals.")

# CSV column -> display label, in the order the Registrar General reports them.
RELIGION_COLUMNS = {
    "Hindus": "Hindu",
    "Muslims": "Muslim",
    "Christians": "Christian",
    "Sikhs": "Sikh",
    "Buddhists": "Buddhist",
    "Jains": "Jain",
    "Others_Religions": "Other religions",
    "Religion_Not_Stated": "Not stated",
}
SCHEDULED_COLUMNS = {"SC": "Scheduled Caste", "ST": "Scheduled Tribe"}

# What the Registrar General published. A mirror that does not reproduce these
# is not the census, whatever it claims to be, so the run stops.
NATIONAL_CONTROLS = {
    "districts": (640, 0),           # (expected, tolerance)
    "population": (1_210_854_977, 0),
    "Hindus": (79.80, 0.05),         # percentages
    "Muslims": (14.23, 0.05),
    "Christians": (2.30, 0.05),
    "Sikhs": (1.72, 0.05),
    "Buddhists": (0.70, 0.05),
    "Jains": (0.37, 0.05),
}

# Districts the 2011 census reported under a name the current boundary files
# spell differently. Renames only -- never a merge or a split.
DISTRICT_ALIASES = {
    "y.s.r.": "Kadapa(YSR)",
    "pondicherry": "Puducherry",
}

# States the 2011 census reported under a different name.
STATE_ALIASES = {
    "orissa": "Odisha",              # renamed 2011
    "pondicherry": "Puducherry",     # renamed 2006
    "nct of delhi": "NCT of Delhi",
}

# States that did not exist at the 2011 census, so no census row covers them.
# Their territory was enumerated under the predecessor state; splitting that
# retrospectively would be an estimate, not a measurement, so they get an
# explicit gap carrying the reason.
FORMED_AFTER_2011 = {
    "Telangana": ("Telangana was formed in 2014 from ten districts of Andhra "
                  "Pradesh. The 2011 census enumerated that territory as part of "
                  "Andhra Pradesh, so no census figure exists for Telangana as "
                  "such. Its districts do carry 2011 figures."),
    "Ladakh": ("Ladakh became a union territory in 2019, split from Jammu and "
               "Kashmir. The 2011 census enumerated it as part of Jammu and "
               "Kashmir, so no census figure exists for Ladakh as such. Its "
               "districts (Leh, Kargil) do carry 2011 figures."),
}

# 2011 districts that have since been subdivided, so one census row covers
# several present-day boundary units. Their figures are deliberately NOT
# spread across the successors: the census never measured those areas
# separately, and inventing a split would be fabrication. The successors keep
# an explicit gap carrying this explanation.
SUBDIVIDED_SINCE_2011 = {
    "jaintia hills": ["East Jaintia Hills", "West Jaintia Hills"],
    "karbi anglong": ["Karbi Anglong East", "Karbi Anglong West"],
    "warangal": ["Warangal (R)", "Warangal (U)"],
}


def cell(row: dict[str, str], key: str) -> int:
    raw = row.get(key)
    if raw in (None, "", "NA", "N/A"):
        return 0
    try:
        return int(float(raw))
    except ValueError:
        return 0


def load_csv(url: str, cache: bool = True) -> list[dict[str, str]]:
    text = http_get(url, cache=cache, timeout=300)
    assert isinstance(text, str)
    return list(csv.DictReader(io.StringIO(text.lstrip("﻿"))))


def validate(rows: list[dict[str, str]]) -> None:
    """Refuse to emit anything unless the extract reproduces the official totals."""
    problems: list[str] = []

    expected, tol = NATIONAL_CONTROLS["districts"]
    if abs(len(rows) - expected) > tol:
        problems.append(f"{len(rows)} districts, expected {expected}")

    total = sum(cell(r, "Population") for r in rows)
    expected, tol = NATIONAL_CONTROLS["population"]
    if abs(total - expected) > tol:
        problems.append(f"population {total:,}, expected {expected:,}")

    for column in RELIGION_COLUMNS:
        control = NATIONAL_CONTROLS.get(column)
        if not control or not total:
            continue
        expected, tol = control
        got = 100.0 * sum(cell(r, column) for r in rows) / total
        if abs(got - expected) > tol:
            problems.append(f"{column} {got:.2f}%, expected {expected:.2f}%")

    if problems:
        raise SystemExit(
            "india_census: the extract does not reproduce the Registrar General's "
            "published Census 2011 totals, so it is not being used:\n  - "
            + "\n  - ".join(problems)
            + f"\nCheck the source ({CSV_URL}) or download the official C-01 "
              f"workbooks from {CATALOG} and re-run with --input.")

    log(f"  validated against published national totals: {len(rows)} districts, "
        f"{total:,} people")


def build_record(name: str, counts: collections.Counter, *, level: str,
                 parent: str, entity_id: str, codes: dict[str, Any]) -> dict[str, Any]:
    population = counts["Population"]
    religion_counts = {label: counts[col] for col, label in RELIGION_COLUMNS.items()
                       if counts[col]}
    scheduled_counts = {label: counts[col] for col, label in SCHEDULED_COLUMNS.items()
                        if counts[col]}

    males, females = counts["Male"], counts["Female"]
    sex_ratio = round(1000.0 * females / males) if males else None

    return record(
        entity_id, name, level=level, parent=parent, codes=codes,
        population=(measure(population, year=2011, source=SOURCE)
                    if population else gap(NOT_AVAILABLE)),
        sex_ratio=(measure(sex_ratio, unit="females_per_1000_males",
                           year=2011, source=SOURCE)
                   if sex_ratio else gap(NOT_AVAILABLE)),
        religion=shares(religion_counts, total=population) or gap(NOT_AVAILABLE),
        religion_note=("Census of India 2011 table C-01. India's next census was "
                       "postponed, so these remain the most recent official figures."),
        religion_year=2011,
        scheduled_groups=shares(scheduled_counts, total=population) or gap(NOT_AVAILABLE),
        scheduled_groups_note=(
            "Scheduled Caste and Scheduled Tribe shares (Census 2011). These are "
            "constitutional-schedule classifications used for reservation policy, "
            "not ethnic categories, and the two do not overlap."),
        ethnicity=gap(NOT_COLLECTED,
                      "India does not collect ethnicity. Scheduled Caste / Scheduled "
                      "Tribe status and mother tongue are collected instead."),
        language=gap(NOT_AVAILABLE,
                     "Mother tongue is Census 2011 table C-16, which is not in this "
                     "extract; it is published as per-state workbooks."),
        sources=[{"field": "population/religion/scheduled groups", "name": SOURCE,
                  "url": CATALOG, "year": 2011,
                  "license": "Government of India open data (GODL-India)",
                  "note": SOURCE_NOTE}],
    )


def districts(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    numeric = list(RELIGION_COLUMNS) + list(SCHEDULED_COLUMNS) + ["Population", "Male", "Female"]

    for row in rows:
        name = (row.get("District name") or "").strip()
        if not name:
            continue
        key = name.lower()
        counts = collections.Counter({col: cell(row, col) for col in numeric})
        code = (row.get("District code") or "").strip()

        if key in SUBDIVIDED_SINCE_2011:
            # One census row, several present-day districts: emit the gap, not a guess.
            for successor in SUBDIVIDED_SINCE_2011[key]:
                out.append(record(
                    f"IND-D{code}-{successor}", successor, level="admin2", parent="IND",
                    religion=gap(NOT_AVAILABLE,
                                 f"The 2011 census reported this area as part of the "
                                 f"undivided {name} district, which has since been "
                                 f"subdivided. The census never measured the successor "
                                 f"districts separately."),
                    ethnicity=gap(NOT_COLLECTED,
                                  "India does not collect ethnicity."),
                    sources=[{"field": "note", "name": SOURCE, "url": CATALOG}],
                ))
            continue

        out.append(build_record(
            DISTRICT_ALIASES.get(key, name), counts,
            level="admin2", parent="IND", entity_id=f"IND-D{code}",
            codes={"census2011_district": code,
                   "state_name": (row.get("State name") or "").strip()}))
    return out


def states(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    """State totals summed from the district rows.

    Plain arithmetic on official counts, not estimation: the sums reproduce the
    published state populations exactly.
    """
    numeric = list(RELIGION_COLUMNS) + list(SCHEDULED_COLUMNS) + ["Population", "Male", "Female"]
    agg: dict[str, collections.Counter] = collections.defaultdict(collections.Counter)
    for row in rows:
        state = (row.get("State name") or "").strip()
        if not state:
            continue
        for col in numeric:
            agg[state][col] += cell(row, col)

    out = []
    for state, counts in sorted(agg.items()):
        # The extract shouts state names; the boundary files use title case.
        name = state.title().replace(" And ", " and ").replace(" Of ", " of ")
        name = STATE_ALIASES.get(state.lower(), name)
        out.append(build_record(
            name, counts, level="admin1", parent="IND",
            entity_id=f"IND-S-{name.replace(' ', '-')}",
            codes={"census2011_state_name": state}))

    for name, reason in FORMED_AFTER_2011.items():
        out.append(record(
            f"IND-S-{name.replace(' ', '-')}", name, level="admin1", parent="IND",
            religion=gap(NOT_AVAILABLE, reason),
            population=gap(NOT_AVAILABLE, reason),
            ethnicity=gap(NOT_COLLECTED, "India does not collect ethnicity."),
            sources=[{"field": "note", "name": SOURCE, "url": CATALOG}],
        ))
    return out


# ---------------------------------------------------------------------------
# Legacy path: official per-state XLSX workbooks
# ---------------------------------------------------------------------------

def read_workbook(path: Path) -> list[dict[str, Any]]:
    """Read one downloaded C-01 workbook.  Requires ``openpyxl`` for .xlsx."""
    if path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover
            raise SystemExit("pip install openpyxl to read censusindia workbooks") from exc
        book = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    else:
        with path.open(newline="", encoding="utf-8-sig") as fh:
            rows = [row for row in csv.reader(fh)]

    header_idx = next((i for i, row in enumerate(rows)
                       if any(isinstance(c, str) and "hindu" in c.lower() for c in row)), None)
    if header_idx is None:
        return []
    header = [str(c or "").strip().lower() for c in rows[header_idx]]
    return [dict(zip(header, row)) for row in rows[header_idx + 1:] if any(row)]


def from_workbooks(folder: Path, level: str) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    columns = {"hindu": "Hindus", "muslim": "Muslims", "christian": "Christians",
               "sikh": "Sikhs", "buddhist": "Buddhists", "jain": "Jains"}
    for path in sorted(folder.glob("*")):
        if path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        for row in read_workbook(path):
            name = str(row.get("area name") or row.get("name") or "").strip()
            if not name:
                continue
            counts = collections.Counter()
            for header, column in columns.items():
                value = next((v for k, v in row.items() if k.startswith(header)), None)
                if isinstance(value, (int, float)):
                    counts[column] = int(value)
            counts["Population"] = int(row.get("total population") or sum(counts.values()))
            code = str(row.get("state code") or row.get("district code") or name).strip()
            out.append(build_record(
                name.title(), counts,
                level="admin1" if level == "state" else "admin2",
                parent="IND", entity_id=f"IND-{code}",
                codes={"census2011": code}))
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--level", default="district", choices=["state", "district"])
    ap.add_argument("--csv-url", default=CSV_URL,
                    help="district-level Census 2011 extract to read")
    ap.add_argument("--input", type=Path, default=None,
                    help="directory of official C-01 workbooks; overrides --csv-url")
    ap.add_argument("--list-sources", action="store_true")
    ap.add_argument("--out", type=Path, default=None)
    args = ap.parse_args()

    if args.list_sources:
        print(f"Census of India 2011 NADA catalogue: {CATALOG}")
        print("  C-01  Population by religious community")
        print("  C-16  Population by mother tongue")
        print(f"District-level extract used by default: {args.csv_url}")
        return 0

    if args.input:
        log(f"india_census: reading workbooks from {args.input}")
        if not args.input.exists():
            log("  no such directory; run with --list-sources for the download URLs")
            return 1
        records = from_workbooks(args.input, args.level)
    else:
        log(f"india_census: Census 2011, level={args.level}")
        rows = load_csv(args.csv_url)
        validate(rows)
        records = states(rows) if args.level == "state" else districts(rows)

    out = args.out or PROCESSED / f"india_{args.level}.json"
    write_json(out, records)
    with_religion = sum(1 for r in records if isinstance(r.get("religion"), list))
    log(f"  {len(records)} records, {with_religion} with religion shares")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
